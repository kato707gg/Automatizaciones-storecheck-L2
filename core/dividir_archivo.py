"""
Lógica de división de archivos Excel en partes.
Sin dependencias de Qt — se puede llamar desde CLI o desde la UI.
"""

import os
import math
import shutil
import openpyxl


def dividir_archivo(
    ruta_archivo: str,
    nombre_hoja: str,
    max_filas: int,
    carpeta_salida: str,
    progreso_cb=None,
) -> int:
    """
    Divide el archivo Excel en partes de max_filas filas cada una.
    Cada parte conserva el encabezado original.
    El archivo fuente no se modifica.

    Parámetros
    ----------
    ruta_archivo   : ruta al .xlsx / .xlsm / .xls de origen.
    nombre_hoja    : nombre exacto de la hoja a dividir.
    max_filas      : límite de filas de datos por archivo (sin contar encabezado).
    carpeta_salida : carpeta donde se guardan las partes generadas.
    progreso_cb    : callable(parte_actual: int, total_partes: int) opcional.

    Retorna
    -------
    Número de partes creadas.

    Excepciones
    -----------
    ValueError  si la hoja no existe, está vacía o no hace falta dividir.
    """
    wb = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)

    if nombre_hoja not in wb.sheetnames:
        hojas = ", ".join(wb.sheetnames)
        raise ValueError(
            f"No se encontró la hoja '{nombre_hoja}'.\n"
            f"Hojas disponibles: {hojas}")

    ws = wb[nombre_hoja]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()

    if not filas:
        raise ValueError("La hoja está vacía.")

    encabezado  = list(filas[0])
    datos       = filas[1:]
    total_filas = len(datos)

    if total_filas <= max_filas:
        raise ValueError(
            f"El archivo tiene {total_filas:,} filas de datos, "
            f"menor o igual al límite de {max_filas:,}. "
            f"No es necesario dividir.")

    nombre_base  = os.path.splitext(os.path.basename(ruta_archivo))[0]
    extension    = os.path.splitext(ruta_archivo)[1]
    total_partes = math.ceil(total_filas / max_filas)

    for i in range(total_partes):
        chunk = datos[i * max_filas: min((i + 1) * max_filas, total_filas)]

        nombre_parte = f"{nombre_base}_Parte{i + 1}{extension}"
        ruta_parte = os.path.join(carpeta_salida, nombre_parte)
        # Copiar el archivo original para preservar formato, estilos y propiedades
        shutil.copy2(ruta_archivo, ruta_parte)
        wb_nuevo = openpyxl.load_workbook(ruta_parte, keep_links=False)
        ws_nuevo = wb_nuevo[nombre_hoja]
        # Eliminar todas las filas de datos (conservar sólo el encabezado en fila 1)
        filas_en_copia = ws_nuevo.max_row
        if filas_en_copia > 1:
            ws_nuevo.delete_rows(2, filas_en_copia - 1)
        for fila in chunk:
            ws_nuevo.append(list(fila))
        wb_nuevo.save(ruta_parte)
        wb_nuevo.close()

        if progreso_cb:
            progreso_cb(i + 1, total_partes)

    return total_partes
