import os
from collections import defaultdict
from datetime import datetime, timedelta
from copy import copy
import shutil
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def catalogacion_solo_por_formato(ruta_matriz, ruta_base, carpeta_trabajo):
    """
    Procesa la catalogación por formato.
    
    Args:
        ruta_matriz: Ruta al archivo MATRIZ DE CATALOGACIÓN en la subcarpeta
        ruta_base: Ruta a la carpeta principal (donde está layout_format_scope_188865)
        carpeta_trabajo: Ruta a la subcarpeta de trabajo
    """
    print("\n" + "="*50)
    print("--- Catalogación Solo Por Formato ---")
    print("="*50)
    
    # Verificar que el archivo matriz existe
    if not os.path.exists(ruta_matriz):
        print(f"  ✗ ERROR: No se encontró el archivo matriz: {ruta_matriz}")
        return False
    
    # Buscar archivo layout_format_scope_188865 en la carpeta base
    archivo_plantilla = None
    for archivo in os.listdir(ruta_base):
        if "layout_format_scope_188865" in archivo.lower() and archivo.endswith(".xlsx"):
            archivo_plantilla = os.path.join(ruta_base, archivo)
            break
    
    if not archivo_plantilla or not os.path.exists(archivo_plantilla):
        print(f"  ✗ ERROR: No se encontró layout_format_scope_188865.xlsx en {ruta_base}")
        return False
    
    print(f"  Archivo matriz: {os.path.basename(ruta_matriz)}")
    print(f"  Archivo plantilla: {os.path.basename(archivo_plantilla)}")
    
    # Parámetros fijos para leer la matriz
    nombre_hoja = "CONFIGURACIÓN DE ANAQUEL"
    fila_id_productos = 3
    fila_encabezados_productos = 4
    fila_inicio = 5
    col_id_formato = 6      # F
    col_nombre_formato = 7  # G
    col_inicio_productos = 11  # K
    
    print(f"\n  Procesando hoja: {nombre_hoja}")
    print("  Cargando datos...")
    
    try:
        # Abrir en modo read_only + data_only para carga más rápida
        wb_matriz = openpyxl.load_workbook(ruta_matriz, data_only=True, read_only=True)
        
        if nombre_hoja not in wb_matriz.sheetnames:
            print(f"  ✗ ERROR: No se encontró la hoja '{nombre_hoja}'")
            wb_matriz.close()
            return False
        
        ws_matriz = wb_matriz[nombre_hoja]
        
        # Pre-cargar toda la hoja en una lista de listas para acceso ultrarrápido
        print("  Pre-cargando matriz en memoria...")
        matriz_rows = [list(row) for row in ws_matriz.iter_rows(values_only=True)]
        wb_matriz.close()
        
        ultima_fila = len(matriz_rows)
        ultima_columna = max((len(r) for r in matriz_rows), default=0)
        
        print(f"  Dimensiones: {ultima_fila} filas x {ultima_columna} columnas")
        
        def gc(row, col):
            """Acceso rápido a celda (índices 1-based) desde lista en memoria."""
            try:
                return matriz_rows[row - 1][col - 1]
            except IndexError:
                return None
        
        # Diccionarios para contar y almacenar datos
        dict_productos_por_formato = defaultdict(int)
        dict_conteo_formatos = defaultdict(int)
        dict_ya_agregado = set()
        
        # Cache de la fila de IDs y nombres de productos para no releerla en cada iteración
        fila_ids = matriz_rows[fila_id_productos - 1] if fila_id_productos <= ultima_fila else []
        fila_nombres = matriz_rows[fila_encabezados_productos - 1] if fila_encabezados_productos <= ultima_fila else []
        
        print("  Paso 1: Contando ocurrencias...")
        
        # Paso 1: Contar ocurrencias (un solo pase sobre datos en memoria)
        for i in range(fila_inicio, ultima_fila + 1):
            try:
                id_formato = gc(i, col_id_formato)
                nombre_formato = gc(i, col_nombre_formato)
                
                if id_formato is not None and nombre_formato is not None:
                    id_formato_str = str(id_formato).strip()
                    
                    if id_formato_str and id_formato_str != 'nan':
                        dict_conteo_formatos[id_formato_str] += 1
                        
                        fila_datos = matriz_rows[i - 1]
                        for j in range(col_inicio_productos - 1, min(ultima_columna, len(fila_datos))):
                            if fila_datos[j] == 1:
                                try:
                                    id_producto = fila_ids[j]
                                except IndexError:
                                    id_producto = None
                                if id_producto is not None:
                                    id_producto_str = str(id_producto).strip()
                                    if id_producto_str and id_producto_str != 'nan':
                                        clave = f"{id_formato_str}|{id_producto_str}"
                                        dict_productos_por_formato[clave] += 1
            except Exception:
                continue
        
        print(f"  Formatos únicos encontrados: {len(dict_conteo_formatos)}")
        print(f"  Combinaciones formato-producto: {len(dict_productos_por_formato)}")
        
        # Preparar datos de salida
        datos_salida = []
        
        print("  Paso 2: Generando datos de catalogación...")
        
        # Paso 2: Crear datos para el archivo de formato (datos ya en memoria)
        for i in range(fila_inicio, ultima_fila + 1):
            try:
                id_formato = gc(i, col_id_formato)
                nombre_formato = gc(i, col_nombre_formato)
                
                if id_formato is not None and nombre_formato is not None:
                    id_formato_str = str(id_formato).strip()
                    
                    if id_formato_str and id_formato_str != 'nan':
                        fila_datos = matriz_rows[i - 1]
                        for j in range(col_inicio_productos - 1, min(ultima_columna, len(fila_datos))):
                            if fila_datos[j] == 1:
                                try:
                                    id_producto = fila_ids[j]
                                    nombre_producto = fila_nombres[j] if j < len(fila_nombres) else None
                                except IndexError:
                                    id_producto = None
                                    nombre_producto = None
                                
                                if id_producto is not None:
                                    id_producto_str = str(id_producto).strip()
                                    clave = f"{id_formato_str}|{id_producto_str}"
                                    
                                    if (clave in dict_productos_por_formato and 
                                        dict_productos_por_formato[clave] == dict_conteo_formatos[id_formato_str] and
                                        clave not in dict_ya_agregado):
                                        
                                        datos_salida.append({
                                            'ID_Formato': id_formato,
                                            'Nombre_Formato': nombre_formato,
                                            'ID_Producto': id_producto,
                                            'Nombre_Producto': nombre_producto
                                        })
                                        
                                        dict_ya_agregado.add(clave)
            except Exception:
                continue
        
        print(f"  Registros a escribir: {len(datos_salida)}")
        
        if not datos_salida:
            print("  ! No hay datos para escribir")
            return True
        
        # Abrir archivo plantilla (sin modificar el original)
        print(f"\n  Abriendo plantilla: {os.path.basename(archivo_plantilla)}")
        wb_plantilla = openpyxl.load_workbook(archivo_plantilla)
        
        hoja_destino = "ProductosCatalogados_Formato"
        if hoja_destino not in wb_plantilla.sheetnames:
            print(f"  ✗ ERROR: No se encontró la hoja '{hoja_destino}'")
            wb_plantilla.close()
            return False
        
        ws_destino = wb_plantilla[hoja_destino]
        
        # Crear una copia de la plantilla con fecha/timestamp para no sobrescribir el original
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_base = os.path.basename(archivo_plantilla)
        nombre_sin_extension = os.path.splitext(nombre_base)[0]
        nombre_archivo_salida = f"{nombre_sin_extension}_catalogado_{timestamp}.xlsx"
        # Guardar directamente en la subcarpeta de trabajo (no en la carpeta base de la plantilla)
        ruta_archivo_salida = os.path.join(carpeta_trabajo, nombre_archivo_salida)
        
        # Obtener estilo de la fila 2 para mantener formato
        estilos_fila2 = {}
        for col in range(1, 7):  # Columnas A-F
            celda_ref = ws_destino.cell(row=2, column=col)
            estilos_fila2[col] = {
                'font': copy(celda_ref.font) if celda_ref.font else None,
                'fill': copy(celda_ref.fill) if celda_ref.fill else None,
                'border': copy(celda_ref.border) if celda_ref.border else None,
                'alignment': copy(celda_ref.alignment) if celda_ref.alignment else None,
                'number_format': celda_ref.number_format
            }
        
        # Calcular fecha de mañana (usar año completo dd/mm/aaaa)
        fecha_manana = datetime.now() + timedelta(days=1)
        fecha_date = fecha_manana.date()
        fecha_str = fecha_date.strftime("%d/%m/%Y")
        
        print(f"  Fecha a colocar: {fecha_str}")
        print("  Escribiendo datos...")
        
        # Escribir datos desde fila 2
        fila_actual = 2
        for dato in datos_salida:
            # Columna A: ID Formato
            celda_a = ws_destino.cell(row=fila_actual, column=1)
            celda_a.value = dato['ID_Formato']
            aplicar_estilo(celda_a, estilos_fila2.get(1))
            
            # Columna B: Nombre Formato
            celda_b = ws_destino.cell(row=fila_actual, column=2)
            celda_b.value = dato['Nombre_Formato']
            aplicar_estilo(celda_b, estilos_fila2.get(2))
            
            # Columna C: ID Producto
            celda_c = ws_destino.cell(row=fila_actual, column=3)
            celda_c.value = dato['ID_Producto']
            aplicar_estilo(celda_c, estilos_fila2.get(3))
            
            # Columna D: Nombre Producto
            celda_d = ws_destino.cell(row=fila_actual, column=4)
            celda_d.value = dato['Nombre_Producto']
            aplicar_estilo(celda_d, estilos_fila2.get(4))
            
            # Columna E: Fecha de mañana (escribir como fecha con formato dd/mm/aaaa)
            celda_e = ws_destino.cell(row=fila_actual, column=5)
            celda_e.value = fecha_date
            aplicar_estilo(celda_e, estilos_fila2.get(5))
            try:
                celda_e.number_format = "dd/mm/yyyy"
            except Exception:
                pass
            
            # Columna F: "INSERT"
            celda_f = ws_destino.cell(row=fila_actual, column=6)
            celda_f.value = "INSERT"
            aplicar_estilo(celda_f, estilos_fila2.get(6))
            
            fila_actual += 1
        
        # Guardar archivo con nombre nuevo (preservar la plantilla original intacta)
        wb_plantilla.save(ruta_archivo_salida)
        wb_plantilla.close()
        print(f"  ✓ Archivo catalogado guardado en subcarpeta: {nombre_archivo_salida}")
        print(f"    (La plantilla original se mantiene intacta y no se escribe en la carpeta base)")
        
        print(f"\n  ¡Catalogación por formato completada!")
        print(f"  Registros agregados: {len(datos_salida)}")
        
        return True
        
    except Exception as e:
        print(f"  ✗ ERROR al procesar: {e}")
        import traceback
        traceback.print_exc()
        return False


def aplicar_estilo(celda, estilo):
    """Aplica un diccionario de estilos a una celda."""
    if not estilo:
        return
    try:
        if estilo.get('font'):
            celda.font = estilo['font']
        if estilo.get('fill'):
            celda.fill = estilo['fill']
        if estilo.get('border'):
            celda.border = estilo['border']
        if estilo.get('alignment'):
            celda.alignment = estilo['alignment']
        if estilo.get('number_format'):
            celda.number_format = estilo['number_format']
    except Exception:
        pass


# Para compatibilidad si se ejecuta directamente
