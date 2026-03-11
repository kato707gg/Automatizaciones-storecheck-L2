"""
Script completo de catalogación.

Este script:
1. Permite elegir una subcarpeta de trabajo
2. Verifica que existan los archivos necesarios
3. Procesa el archivo MATRIZ DE CATALOGACIÓN
4. Ejecuta catalogación por formato
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy
import os

# Importar módulo de catalogación por formato
from .formato import catalogacion_solo_por_formato
from .tienda import catalogacion_por_tienda
from .filtro_formato import rellenar_unos_configuracion_anaquel, procesar_filtro_formato
from .producto_general import procesar_producto_general




# ---------------------------------------------------------------------------
# Utilidades de carga rápida
# ---------------------------------------------------------------------------

def _leer_wb_hojas(ruta, *nombres_hojas):
    """Abre un Excel UNA sola vez en modo lectura y devuelve un dict
    {nombre_hoja: list[list]} con los datos de las hojas solicitadas."""
    result = {n: [] for n in nombres_hojas}
    if not os.path.exists(ruta):
        return result
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
        for nombre in nombres_hojas:
            if nombre in wb.sheetnames:
                ws = wb[nombre]
                result[nombre] = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
    except Exception:
        pass
    return result


def _gv(rows, row_1based, col_1based):
    """Acceso seguro (índices 1-based) a una lista-de-listas leída en memoria."""
    try:
        return rows[row_1based - 1][col_1based - 1]
    except IndexError:
        return None



def copiar_hoja(wb, nombre_origen, nombre_destino):
    """Copia una hoja del libro de Excel."""
    if nombre_origen not in wb.sheetnames:
        print(f"  ✗ ERROR: No se encontró la hoja '{nombre_origen}'")
        return False
    
    # Verificar si ya existe la hoja destino
    if nombre_destino in wb.sheetnames:
        print(f"  ! La hoja '{nombre_destino}' ya existe, se omite la copia.")
        return True
    
    # Copiar la hoja
    hoja_origen = wb[nombre_origen]
    hoja_copia = wb.copy_worksheet(hoja_origen)
    hoja_copia.title = nombre_destino
    
    print(f"  ✓ Hoja copiada: '{nombre_origen}' -> '{nombre_destino}'")
    return True


def crear_hoja(wb, nombre_hoja):
    """Crea una nueva hoja en el libro de Excel."""
    if nombre_hoja in wb.sheetnames:
        print(f"  ! La hoja '{nombre_hoja}' ya existe, se omite la creación.")
        return wb[nombre_hoja]
    
    nueva_hoja = wb.create_sheet(title=nombre_hoja)
    print(f"  ✓ Hoja creada: '{nombre_hoja}'")
    return nueva_hoja


def limpiar_hoja(ws):
    """Elimina todo el contenido de la hoja dada."""
    for fila in ws.iter_rows():
        for celda in fila:
            celda.value = None
    print(f"  ✓ Contenido limpiado en hoja '{ws.title}'")


def copiar_celda_con_estilo(src_cell, dst_cell):
    """Copia valor y estilo de una celda a otra."""
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)


def normalizar_columna_c_productos(ws, fila_inicio):
    """Convierte la columna C a número cuando sea posible (equivalente a Datos > Texto en columnas)."""
    col_c = column_index_from_string('C')
    celdas = 0
    for fila in range(fila_inicio, ws.max_row + 1):
        celda = ws.cell(row=fila, column=col_c)
        val = celda.value
        if val is None:
            continue
        texto = str(val).strip()
        if texto == "":
            continue
        # Intentar convertir a entero o float
        try:
            if texto.isdigit() or (texto.startswith('-') and texto[1:].isdigit()):
                celda.value = int(texto)
            else:
                celda.value = float(texto)
        except ValueError:
            celda.value = texto
            celda.number_format = "@"
        celdas += 1
    print(f"  ✓ Columna C normalizada (texto a número si aplica) desde fila {fila_inicio} ({celdas} celdas)")


def copiar_configuracion_anaquel(wb):
    """Copia filas 2-4 desde columna G en CONFIGURACIÓN DE ANAQUEL 1 hacia PRODUCTOS (fila 1 col E)."""
    print("\n--- Copiando CONFIGURACIÓN DE ANAQUEL 1 a PRODUCTOS ---")

    if "CONFIGURACIÓN DE ANAQUEL 1" not in wb.sheetnames:
        print("  ✗ ERROR: No existe la hoja 'CONFIGURACIÓN DE ANAQUEL 1'.")
        return False, 0, 0
    if "PRODUCTOS" not in wb.sheetnames:
        print("  ✗ ERROR: No existe la hoja 'PRODUCTOS' en el archivo destino.")
        return False, 0, 0

    ws_src = wb["CONFIGURACIÓN DE ANAQUEL 1"]
    ws_dst = wb["PRODUCTOS"]

    fila_origenes = [2, 3, 4]
    col_inicio = column_index_from_string('G')

    # Determinar última columna con datos en filas 2-4 desde G en adelante (permitiendo huecos)
    max_col = col_inicio - 1
    for col in range(col_inicio, ws_src.max_column + 1):
        if any(ws_src.cell(row=f, column=col).value not in (None, "") for f in fila_origenes):
            max_col = col

    if max_col < col_inicio:
        print("  ! No se encontraron datos en filas 2-4 desde la columna G.")
        return True, 0, column_index_from_string('E')

    ancho = max_col - col_inicio + 1
    col_dest_inicio = column_index_from_string('E')  # columna E

    # Copiar filas 2-4 -> destino filas 1-3 (con estilos)
    for idx, fila_src in enumerate(fila_origenes):
        fila_dst = 1 + idx
        for offset in range(ancho):
            src_cell = ws_src.cell(row=fila_src, column=col_inicio + offset)
            dst_cell = ws_dst.cell(row=fila_dst, column=col_dest_inicio + offset)
            copiar_celda_con_estilo(src_cell, dst_cell)

    # Copiar fila 1 (encabezado copiado) hacia fila 4, mismas columnas
    fila_origen_enc = 1
    fila_dst_enc = 4
    for offset in range(ancho):
        src_cell = ws_dst.cell(row=fila_origen_enc, column=col_dest_inicio + offset)
        dst_cell = ws_dst.cell(row=fila_dst_enc, column=col_dest_inicio + offset)
        copiar_celda_con_estilo(src_cell, dst_cell)

    # Aplicar estilo a filas 5 y 6 en el rango copiado
    try:
        aplicar_formato_filas_5_6(ws_dst, col_dest_inicio, ancho)
        aplicar_color_fila5(ws_dst, col_dest_inicio, ancho, color_hex="00B0F0")
    except Exception:
        pass

    print(f"  ✓ Copiado rango G2:{get_column_letter(max_col)}4 -> PRODUCTOS!E1 con ancho {ancho} columnas")
    print("  ✓ Fila 1 replicada en fila 4 en columnas E en adelante")
    return True, ancho, col_dest_inicio


def copiar_productos(ruta_layout, wb_destino):
    """Copia columnas A, B y E desde layout_products a la hoja PRODUCTOS (cols A-C) iniciando en fila 8."""
    print("\n--- Copiando productos ---")

    if "PRODUCTOS" not in wb_destino.sheetnames:
        print("  ✗ ERROR: No existe la hoja 'PRODUCTOS' en el archivo destino.")
        return False, 0

    if not os.path.exists(ruta_layout):
        print(f"  ✗ ERROR: No se encontró el layout de productos: {ruta_layout}")
        return False, 0

    wb_layout = openpyxl.load_workbook(ruta_layout, data_only=True, read_only=True)

    if "Productos" not in wb_layout.sheetnames:
        print("  ✗ ERROR: No se encontró la hoja 'Productos' en layout_products.")
        wb_layout.close()
        return False, 0

    ws_src = wb_layout["Productos"]

    # Pre-cargar hoja en memoria para acceso rápido
    print("  Pre-cargando layout de productos en memoria...")
    rows_src = [list(r) for r in ws_src.iter_rows(values_only=True)]
    wb_layout.close()

    ws_dst = wb_destino["PRODUCTOS"]

    fila_encabezado = 7
    fila_destino = 8
    filas_copiadas = 0
    columnas_origen = [1, 2, 5]  # A, B, E
    col_filtro_q = 17            # Columna Q (índice 0-based: 16)

    # Copiar encabezados de A,B,E en fila 7 (fila 1 de layout = rows_src[0])
    if rows_src:
        enc_row = rows_src[0]
        for idx, col in enumerate(columnas_origen, start=1):
            val = enc_row[col - 1] if col - 1 < len(enc_row) else None
            ws_dst.cell(row=fila_encabezado, column=idx).value = val

    # Copiar datos filtrando por Q==1, desde fila 2 (rows_src[1:])
    for fila_data in rows_src[1:]:
        valor_q = fila_data[col_filtro_q - 1] if col_filtro_q - 1 < len(fila_data) else None

        # Solo procesar filas donde Q == 1
        if not (valor_q == 1 or str(valor_q).strip() == "1"):
            continue

        valores = [fila_data[col - 1] if col - 1 < len(fila_data) else None for col in columnas_origen]

        # Omitir filas vacías
        if all(v is None or str(v).strip() == "" for v in valores):
            continue

        ws_dst.cell(row=fila_destino, column=1).value = valores[0]
        ws_dst.cell(row=fila_destino, column=2).value = valores[1]
        ws_dst.cell(row=fila_destino, column=3).value = valores[2]

        fila_destino += 1
        filas_copiadas += 1

    # Normalizar columna C como texto para que las búsquedas coincidan
    normalizar_columna_c_productos(ws_dst, fila_inicio=8)

    # Dar formato a encabezados A7:C7
    aplicar_formato_encabezados_productos(ws_dst)

    print(f"  ✓ Filas copiadas: {filas_copiadas}")
    return True, filas_copiadas


def aplicar_formato_encabezados_productos(ws):
    """Aplica formato de encabezado a A7:C7 similar al ejemplo (relleno verde, negritas, centrado)."""
    fill = PatternFill(fill_type="solid", fgColor="9BCB7F")
    font = Font(bold=True, color="000000")
    align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="9E9E9E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, 4):  # A-C
        celda = ws.cell(row=7, column=col)
        celda.fill = fill
        celda.font = font
        celda.alignment = align
        celda.border = border


def aplicar_formato_filas_5_6(ws, col_inicio, ancho):
    """Aplica formato a las filas 5 y 6 en el rango desde col_inicio por ancho columnas.
    Ambas filas: fondo azul claro, texto negro, centrado vertical y horizontal, borde fino.
    """
    fill_azul = PatternFill(fill_type="solid", fgColor="BDD7EE")
    font_black = Font(bold=False, color="000000")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for offset in range(ancho):
        c5 = ws.cell(row=5, column=col_inicio + offset)
        c6 = ws.cell(row=6, column=col_inicio + offset)
        c5.fill = fill_azul
        c5.font = font_black
        c5.alignment = align
        c5.border = border

        c6.fill = fill_azul
        c6.font = font_black
        c6.alignment = align
        c6.border = border


def aplicar_color_fila5(ws, col_inicio, ancho, color_hex="00B0F0"):
    """Aplica un color específico a la fila 5 desde col_inicio por ancho columnas."""
    fill = PatternFill(fill_type="solid", fgColor=color_hex)
    font_black = Font(bold=False, color="000000")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for offset in range(ancho):
        c5 = ws.cell(row=5, column=col_inicio + offset)
        c5.fill = fill
        c5.font = font_black
        c5.alignment = align
        c5.border = border


def eliminar_columnas_dinamico(ws, col_inicio):
    """Elimina dinámicamente columnas de productos discontinuados y placeholders de competencia.

    Reglas (se revisan filas 2 y 3 de PRODUCTOS):
    · Contiene '300ml'       → elimina esa columna Y la siguiente a su derecha
                               (reduce las 2 trailing empties del grupo a 1).
      Excepción: si la columna también contiene 'pediatrico' en cualquiera de
                 las filas 2 o 3, NO se elimina (es un producto vigente).
    · Empieza con 'total '   → elimina sólo esa columna
                               (son placeholders de competencia que se sustituirán
                               con la lista real de 'Productos de Sams').
    Elimina siempre de derecha a izquierda para no alterar índices.
    """
    print("\n--- Eliminando columnas dinámicamente (300ml + competencia) ---")

    cols_300ml = []   # col_idx donde aparece '300ml' → borrar col + col+1
    cols_total = []   # col_idx donde aparece 'total ...' → borrar sólo esa col

    for col in range(col_inicio, ws.max_column + 1):
        # Recopilar todos los textos de las filas 2 y 3 para esta columna
        textos = []
        for fila in (2, 3):
            val = ws.cell(row=fila, column=col).value
            if val is not None:
                textos.append(str(val).strip().lower())

        tiene_300ml    = any("300ml"    in t for t in textos)
        tiene_pediatrico = any("pediatrico" in t for t in textos)
        tiene_total    = any(t.startswith("total ") or t == "total" for t in textos)

        if tiene_300ml and not tiene_pediatrico:
            cols_300ml.append(col)
        elif tiene_total:
            cols_total.append(col)

    to_delete = set()
    for col in cols_300ml:
        to_delete.add(col)
        to_delete.add(col + 1)   # columna extra a la derecha
    for col in cols_total:
        to_delete.add(col)

    for col in sorted(to_delete, reverse=True):
        ws.delete_cols(col)

    print(f"  ✓ Eliminadas {len(to_delete)} columnas en total")
    if cols_300ml:
        print(f"      300ml (+extra): {[get_column_letter(c) for c in sorted(cols_300ml)]}")
    if cols_total:
        print(f"      Total (competencia): {[get_column_letter(c) for c in sorted(cols_total)]}")

def establecer_formulas_id_nombre(ws, col_inicio):
    """Escribe fórmulas de filas 5 (ID) y 6 (Nombre) en columnas con código en fila 2."""
    fila_id = 5
    fila_nombre = 6

    # Detectar última columna con datos en filas 1-3 (sin depender de ancho externo)
    ultima_col = col_inicio
    for col in range(col_inicio, ws.max_column + 1):
        for fila in (1, 2, 3):
            val = ws.cell(row=fila, column=col).value
            if val is not None and str(val).strip():
                ultima_col = col
                break

    for col_idx in range(col_inicio, ultima_col + 1):
        col_letter = get_column_letter(col_idx)

        celda_id = ws.cell(row=fila_id, column=col_idx)
        if celda_id.value is None:
            celda_id.value = f"=_xlfn.XLOOKUP({col_letter}2,$C$8:$C$1000,$A$8:$A$1000)"

        celda_nombre = ws.cell(row=fila_nombre, column=col_idx)
        if celda_nombre.value is None:
            celda_nombre.value = f"=_xlfn.XLOOKUP({col_letter}2,$C$8:$C$1000,$B$8:$B$1000)"


def llenar_espacios_vacios_productos_sams(ws, col_inicio, ruta_base):
    """Lee 'Productos de Sams' y coloca los productos dinámicamente:
      · Generales     → trailing empties de cada grupo (asignación global Jaccard)
      · Sobrantes     → después de todos los grupos
      · Competencia   → después de los sobrantes
      · SAMS          → después de la competencia
    Los grupos y rangos se detectan automáticamente desde la hoja PRODUCTOS.
    No se insertan columnas para no romper la alineación con CONFIG DE ANAQUEL.
    """
    print("\n--- Llenando espacios vacíos con Productos de Sams ---")

    # ── Localizar archivo ────────────────────────────────────────────────────────
    archivo_sams = None
    for archivo in os.listdir(ruta_base):
        if "productos de sams" in archivo.lower():
            archivo_sams = os.path.join(ruta_base, archivo)
            break

    if not archivo_sams:
        print("  ! No se encontró el archivo 'Productos de Sams', se omite este paso.")
        return

    print(f"  ✓ Encontrado: {os.path.basename(archivo_sams)}")

    # ── Leer todo el archivo en memoria ─────────────────────────────────────────
    wb_sams = openpyxl.load_workbook(archivo_sams, data_only=True, read_only=True)
    rows_all = [list(fila) for fila in wb_sams.active.iter_rows(values_only=True)]
    wb_sams.close()

    if not rows_all:
        print("  ! Archivo vacío.")
        return

    header_row = rows_all[0]

    # ── Detectar columnas por texto del encabezado (fila 1) ─────────────────────
    KWORD_NORMAL = "captura normal"
    KWORD_COMP   = "competencia"
    KWORD_SAMS   = "sams"

    col_id_normal = col_nom_normal = None
    col_id_comp   = col_nom_comp   = None
    col_id_sams   = col_nom_sams   = None

    for idx, header in enumerate(header_row):
        if header is None:
            continue
        h = str(header).lower()
        if KWORD_NORMAL in h:
            col_nom_normal = idx
            col_id_normal  = idx - 1
        elif KWORD_COMP in h:
            col_nom_comp = idx
            col_id_comp  = idx - 1
        elif KWORD_SAMS in h:
            col_nom_sams = idx
            col_id_sams  = idx - 1

    def _extraer_pares(col_id, col_nom):
        resultado = []
        for fila in rows_all[1:]:
            id_val  = fila[col_id]  if col_id  < len(fila) else None
            nom_val = fila[col_nom] if col_nom < len(fila) else None
            if id_val is not None:
                resultado.append((id_val, nom_val))
        return resultado

    def _extraer_pares_sams(col_id, col_nom):
        resultado = []
        for fila in rows_all[1:]:
            id_val  = fila[col_id]  if col_id  < len(fila) else None
            nom_val = fila[col_nom] if col_nom < len(fila) else None
            if id_val is not None or nom_val is not None:
                resultado.append((id_val, nom_val))
        return resultado

    productos_normal  = _extraer_pares(col_id_normal, col_nom_normal)  if col_id_normal is not None else []
    productos_comp    = _extraer_pares(col_id_comp,   col_nom_comp)    if col_id_comp   is not None else []
    productos_sams_ef = _extraer_pares_sams(col_id_sams, col_nom_sams) if col_id_sams   is not None else []

    print(f"  ✓ Productos normales:     {len(productos_normal)}")
    print(f"  ✓ Productos competencia:  {len(productos_comp)}")
    print(f"  ✓ Productos SAMS:         {len(productos_sams_ef)}")

    if not productos_normal and not productos_comp and not productos_sams_ef:
        print("  ! No se encontraron productos en ningún grupo.")
        return

    # ── Estilos comunes ──────────────────────────────────────────────────────────
    fill_azul  = PatternFill(fill_type="solid", fgColor="BDD7EE")
    font_black = Font(bold=False, color="000000")
    align      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin       = Side(style="thin", color="000000")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _escribir(fila, col, valor):
        c = ws.cell(row=fila, column=col)
        c.value     = valor
        c.fill      = fill_azul
        c.font      = font_black
        c.alignment = align
        c.border    = border

    # ── Detectar grupos dinámicamente desde fila 1 ───────────────────────────────
    grupos = []
    for col in range(col_inicio, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is not None and str(val).strip() != "":
            grupos.append((col, str(val).strip()))

    if not grupos:
        print("  ! No se detectaron grupos en fila 1.")
        return

    # ── Última columna con datos en fila 1 o fila 2 ─────────────────────────────
    ultima_col_datos = col_inicio
    for col in range(col_inicio, ws.max_column + 1):
        v1 = ws.cell(row=1, column=col).value
        v2 = ws.cell(row=2, column=col).value
        if (v1 is not None and str(v1).strip()) or (v2 is not None and str(v2).strip()):
            ultima_col_datos = col

    print(f"  ✓ Grupos detectados en fila 1: {len(grupos)}")
    for gc, gn in grupos:
        print(f"      {get_column_letter(gc)}({gc}): {gn}")
    print(f"  ✓ Última columna con datos: {get_column_letter(ultima_col_datos)} ({ultima_col_datos})")

    # ── Función de similitud Jaccard ─────────────────────────────────────────────
    def _jaccard(grupo_name, producto_name):
        g_words = set(grupo_name.upper().split())
        p_words = set(str(producto_name).upper().split())
        inter = g_words & p_words
        union = g_words | p_words
        return len(inter) / len(union) if union else 0

    # ── Paso 1: Generales → trailing empties de cada grupo normal ────────────────
    # Recopilar grupos normales con sus trailing empties
    grupos_normales = []  # (g_name, [trailing_cols])
    for i, (g_start, g_name) in enumerate(grupos):
        if "competencia" in g_name.lower():
            continue
        g_end = (grupos[i + 1][0] - 1) if i + 1 < len(grupos) else ultima_col_datos
        trailing = []
        for col in range(g_end, g_start - 1, -1):
            val_f2 = ws.cell(row=2, column=col).value
            if val_f2 is None or str(val_f2).strip() == "":
                trailing.insert(0, col)
            else:
                break
        if trailing:
            grupos_normales.append((g_name, trailing))
        print(f"      {g_name}: rango {get_column_letter(g_start)}-{get_column_letter(g_end)}, "
              f"{len(trailing)} trailing empties")

    # Asignación global por Jaccard CON LÍMITE por grupo (no asignar más que trailing empties)
    pool_normal = list(productos_normal)
    candidatos = []
    for gi, (g_name, _) in enumerate(grupos_normales):
        for pi, (_, nom_p) in enumerate(pool_normal):
            s = _jaccard(g_name, str(nom_p))
            if s > 0:
                candidatos.append((s, gi, pi))
    candidatos.sort(key=lambda x: -x[0])

    asignado = {}       # g_idx -> [p_idx, ...]
    productos_usados = set()
    for score, gi, pi in candidatos:
        if pi in productos_usados:
            continue
        _, trailing = grupos_normales[gi]
        if len(asignado.get(gi, [])) >= len(trailing):
            continue  # No asignar más productos que trailing empties disponibles
        asignado.setdefault(gi, []).append(pi)
        productos_usados.add(pi)

    # Escribir generales asignados
    generales_colocados = 0
    for gi, (g_name, trailing) in enumerate(grupos_normales):
        for j, pi in enumerate(asignado.get(gi, [])):
            id_p, nom_p = pool_normal[pi]
            _escribir(5, trailing[j], id_p)
            _escribir(6, trailing[j], nom_p)
            generales_colocados += 1
            print(f"      {g_name}: {nom_p} → {get_column_letter(trailing[j])}")

    pool_restante = [p for i, p in enumerate(pool_normal) if i not in productos_usados]
    print(f"  ✓ Productos generales colocados: {generales_colocados}")

    # ── Paso 2: Sobrantes normales + Competencia + SAMS → secuenciales ───────────
    col_siguiente = ultima_col_datos + 1

    if pool_restante:
        for id_p, nom_p in pool_restante:
            _escribir(5, col_siguiente, id_p)
            _escribir(6, col_siguiente, nom_p)
            col_siguiente += 1
        print(f"  ✓ Generales sobrantes: {len(pool_restante)} → "
              f"desde {get_column_letter(col_siguiente - len(pool_restante))}")

    if productos_comp:
        col_comp = col_siguiente
        ws.cell(row=4, column=col_comp).value = "Competencia"
        for id_p, nom_p in productos_comp:
            _escribir(5, col_siguiente, id_p)
            _escribir(6, col_siguiente, nom_p)
            col_siguiente += 1
        print(f"  ✓ Competencia: {len(productos_comp)} → desde {get_column_letter(col_comp)}")

    if productos_sams_ef:
        col_sams = col_siguiente
        ws.cell(row=4, column=col_sams).value = "Productos Sams"
        for id_val, nom_val in productos_sams_ef:
            _escribir(5, col_sams, id_val)
            _escribir(6, col_sams, nom_val)
            col_sams += 1
        print(f"  ✓ SAMS: {len(productos_sams_ef)} → desde {get_column_letter(col_siguiente)}")


def procesar_lugares(wb, ruta_layout_places, hojas_lp=None):
    """Procesa la hoja LUGARES: copia datos de layout_places y CONFIGURACIÓN DE ANAQUEL 1."""
    print("\n--- Procesando hoja LUGARES ---")

    if "LUGARES" not in wb.sheetnames:
        print("  ✗ ERROR: No existe la hoja 'LUGARES'.")
        return False

    ws_lugares = wb["LUGARES"]

    # Usar datos pre-cargados si están disponibles; si no, cargar desde archivo
    if hojas_lp is not None:
        rows_layout = hojas_lp.get("Lugares", [])
    else:
        if not os.path.exists(ruta_layout_places):
            print(f"  ✗ ERROR: No se encontró el layout de lugares: {ruta_layout_places}")
            return False
        _tmp = _leer_wb_hojas(ruta_layout_places, "Lugares")
        rows_layout = _tmp.get("Lugares", [])

    if not rows_layout:
        print("  ✗ ERROR: No se encontró la hoja 'Lugares' en layout_places.")
        return False

    # Copiar columnas A, B, F desde layout_places a LUGARES columnas A, B, C
    fila_destino = 1
    filas_copiadas = 0

    for fila_data in rows_layout:
        val_a = fila_data[0] if len(fila_data) > 0 else None
        val_b = fila_data[1] if len(fila_data) > 1 else None
        val_f = fila_data[5] if len(fila_data) > 5 else None
        fila = fila_destino  # para la condición de fila > 1

        # Omitir filas completamente vacías (después del encabezado)
        if fila > 1 and all(v is None or str(v).strip() == "" for v in [val_a, val_b, val_f]):
            continue

        ws_lugares.cell(row=fila_destino, column=1).value = val_a
        ws_lugares.cell(row=fila_destino, column=2).value = val_b

        # Columna F convertida a número (columna C destino)
        if val_f is not None:
            texto = str(val_f).strip()
            try:
                if texto.isdigit() or (texto.startswith('-') and texto[1:].isdigit()):
                    ws_lugares.cell(row=fila_destino, column=3).value = int(texto)
                elif texto:
                    ws_lugares.cell(row=fila_destino, column=3).value = float(texto)
                else:
                    ws_lugares.cell(row=fila_destino, column=3).value = val_f
            except ValueError:
                ws_lugares.cell(row=fila_destino, column=3).value = val_f
        else:
            ws_lugares.cell(row=fila_destino, column=3).value = val_f

        fila_destino += 1
        filas_copiadas += 1

    print(f"  ✓ Copiadas {filas_copiadas} filas de layout_places (cols A,B,F -> A,B,C)")
    print("  ✓ Columna C convertida a número")

    # Copiar columnas D y E desde fila 4 de CONFIGURACIÓN DE ANAQUEL 1 a LUGARES D y E
    max_fila_de = 1
    if "CONFIGURACIÓN DE ANAQUEL 1" in wb.sheetnames:
        ws_config = wb["CONFIGURACIÓN DE ANAQUEL 1"]
        col_d = column_index_from_string('D')  # 4
        col_e = column_index_from_string('E')  # 5

        # Determinar última fila con datos en columnas D o E (desde fila 4)
        max_fila_config = 4
        for fila in range(4, ws_config.max_row + 1):
            val_d = ws_config.cell(row=fila, column=col_d).value
            val_e = ws_config.cell(row=fila, column=col_e).value
            if val_d is not None or val_e is not None:
                max_fila_config = fila

        # Copiar desde fila 4 de CONFIG a fila 1 de LUGARES
        filas_copiadas_config = 0
        for fila_src in range(4, max_fila_config + 1):
            fila_dst = fila_src - 3  # fila 4 -> 1, fila 5 -> 2, etc.
            ws_lugares.cell(row=fila_dst, column=4).value = ws_config.cell(row=fila_src, column=col_d).value
            ws_lugares.cell(row=fila_dst, column=5).value = ws_config.cell(row=fila_src, column=col_e).value
            filas_copiadas_config += 1
            max_fila_de = fila_dst

        print(f"  ✓ Copiado columnas D,E de CONFIGURACIÓN DE ANAQUEL 1 (filas 4-{max_fila_config}) -> LUGARES D,E ({filas_copiadas_config} filas)")
    else:
        print("  ! No se encontró CONFIGURACIÓN DE ANAQUEL 1 para copiar D y E")

    # Aplicar estilos a encabezados (fila 1)
    # A1:C1 verde #7AD694
    fill_verde = PatternFill(fill_type="solid", fgColor="7AD694")
    # D1:E1 amarillo #FFC000
    fill_amarillo = PatternFill(fill_type="solid", fgColor="FFC000")
    # F1:G1 azul #4285F4
    fill_azul = PatternFill(fill_type="solid", fgColor="4285F4")
    align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Aplicar a A1:C1
    for col in range(1, 4):
        celda = ws_lugares.cell(row=1, column=col)
        celda.fill = fill_verde
        celda.alignment = align
        celda.border = border

    # Aplicar a D1:E1
    for col in range(4, 6):
        celda = ws_lugares.cell(row=1, column=col)
        celda.fill = fill_amarillo
        celda.alignment = align
        celda.border = border

    # Poner encabezados en columnas F y G
    ws_lugares.cell(row=1, column=6).value = "ID Sto"
    ws_lugares.cell(row=1, column=7).value = "Nombre Sto"

    # Aplicar a F1:G1
    for col in range(6, 8):
        celda = ws_lugares.cell(row=1, column=col)
        celda.fill = fill_azul
        celda.alignment = align
        celda.border = border

    print("  ✓ Encabezados con estilos aplicados (A-C verde, D-E amarillo, F-G azul)")

    # Agregar fórmulas BUSCARX desde fila 2 hasta la última fila con datos en columnas D y E
    ultima_fila = max_fila_de  # Usar la última fila de D y E
    for fila in range(2, ultima_fila + 1):
        # F: =BUSCARX(D2, $C:$C, $A:$A)
        ws_lugares.cell(row=fila, column=6).value = f"=_xlfn.XLOOKUP(D{fila},$C:$C,$A:$A)"
        # G: =BUSCARX(F2, $A:$A, $B:$B)
        ws_lugares.cell(row=fila, column=7).value = f"=_xlfn.XLOOKUP(F{fila},$A:$A,$B:$B)"

    print(f"  ✓ Fórmulas BUSCARX agregadas en columnas F y G (filas 2-{ultima_fila})")

    return True


def procesar_cadena_canal_formato(wb, ruta_layout_places, hojas_lp=None):
    """Procesa la hoja CADENA CANAL FORMATO: copia hojas ocultas de layout_places."""
    print("\n--- Procesando hoja CADENA CANAL FORMATO ---")

    if "CADENA CANAL FORMATO" not in wb.sheetnames:
        print("  ✗ ERROR: No existe la hoja 'CADENA CANAL FORMATO'.")
        return False

    ws_destino = wb["CADENA CANAL FORMATO"]

    # Usar datos pre-cargados si están disponibles
    if hojas_lp is None:
        if not os.path.exists(ruta_layout_places):
            print(f"  ✗ ERROR: No se encontró el layout de lugares: {ruta_layout_places}")
            return False
        hojas_lp = _leer_wb_hojas(ruta_layout_places, "Formatos", "Cadenas", "Canales")

    # Copiar hoja "Formatos" columnas A,B a columnas A,B
    rows_formatos = hojas_lp.get("Formatos", [])
    if rows_formatos:
        filas_copiadas = 0
        for fi, fila_data in enumerate(rows_formatos, start=1):
            val_a = fila_data[0] if len(fila_data) > 0 else None
            val_b = fila_data[1] if len(fila_data) > 1 else None
            if fi > 1 and (val_a is None or str(val_a).strip() == "") and (val_b is None or str(val_b).strip() == ""):
                continue
            ws_destino.cell(row=fi, column=1).value = val_a
            ws_destino.cell(row=fi, column=2).value = val_b
            filas_copiadas += 1
        print(f"  ✓ Copiadas {filas_copiadas} filas de Formatos (A,B -> A,B)")
    else:
        print("  ! No se encontró la hoja 'Formatos'")

    # Copiar hoja "Cadenas" columnas A,B a columnas C,D
    rows_cadenas = hojas_lp.get("Cadenas", [])
    if rows_cadenas:
        filas_copiadas = 0
        for fi, fila_data in enumerate(rows_cadenas, start=1):
            val_a = fila_data[0] if len(fila_data) > 0 else None
            val_b = fila_data[1] if len(fila_data) > 1 else None
            if fi > 1 and (val_a is None or str(val_a).strip() == "") and (val_b is None or str(val_b).strip() == ""):
                continue
            ws_destino.cell(row=fi, column=3).value = val_a
            ws_destino.cell(row=fi, column=4).value = val_b
            filas_copiadas += 1
        print(f"  ✓ Copiadas {filas_copiadas} filas de Cadenas (A,B -> C,D)")
    else:
        print("  ! No se encontró la hoja 'Cadenas'")

    # Copiar hoja "Canales" columnas A,B a columnas E,F
    rows_canales = hojas_lp.get("Canales", [])
    if rows_canales:
        filas_copiadas = 0
        for fi, fila_data in enumerate(rows_canales, start=1):
            val_a = fila_data[0] if len(fila_data) > 0 else None
            val_b = fila_data[1] if len(fila_data) > 1 else None
            if fi > 1 and (val_a is None or str(val_a).strip() == "") and (val_b is None or str(val_b).strip() == ""):
                continue
            ws_destino.cell(row=fi, column=5).value = val_a
            ws_destino.cell(row=fi, column=6).value = val_b
            filas_copiadas += 1
        print(f"  ✓ Copiadas {filas_copiadas} filas de Canales (A,B -> E,F)")
    else:
        print("  ! No se encontró la hoja 'Canales'")

    # Aplicar estilos a encabezados (fila 1)
    # A1:B1 - Color 1
    fill_formato = PatternFill(fill_type="solid", fgColor="FF6B9D")
    # C1:D1 - Color 2
    fill_cadena = PatternFill(fill_type="solid", fgColor="9BCB7F")
    # E1:F1 - Color 3
    fill_canal = PatternFill(fill_type="solid", fgColor="FFD966")
    
    font_blanco = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Aplicar a A1:B1
    for col in range(1, 3):
        celda = ws_destino.cell(row=1, column=col)
        celda.fill = fill_formato
        celda.font = font_blanco
        celda.alignment = align
        celda.border = border

    # Aplicar a C1:D1
    for col in range(3, 5):
        celda = ws_destino.cell(row=1, column=col)
        celda.fill = fill_cadena
        celda.font = font_blanco
        celda.alignment = align
        celda.border = border

    # Aplicar a E1:F1
    for col in range(5, 7):
        celda = ws_destino.cell(row=1, column=col)
        celda.fill = fill_canal
        celda.font = font_blanco
        celda.alignment = align
        celda.border = border

    print("  ✓ Encabezados con estilos aplicados")

    return True


def procesar_formatos(wb, ruta_layout_places, hojas_lp=None):
    """Procesa la hoja FORMATOS usando datos de layout_places y la hoja LUGARES del libro."""
    print("\n--- Procesando hoja FORMATOS ---")

    if "FORMATOS" not in wb.sheetnames:
        print("  ✗ ERROR: No existe la hoja 'FORMATOS'.")
        return False
    if "LUGARES" not in wb.sheetnames:
        print("  ✗ ERROR: No existe la hoja 'LUGARES' en la matriz para complementar datos.")
        return False

    ws_dst = wb["FORMATOS"]
    ws_lugares = wb["LUGARES"]

    # Usar datos pre-cargados si están disponibles
    if hojas_lp is None:
        if not os.path.exists(ruta_layout_places):
            print(f"  ✗ ERROR: No se encontró el layout de lugares: {ruta_layout_places}")
            return False
        hojas_lp = _leer_wb_hojas(ruta_layout_places, "Formatos", "Lugares")

    # --- Bloque 1: Formatos (layout_places hoja "Formatos") ---
    rows_formatos_lp = hojas_lp.get("Formatos", [])
    if not rows_formatos_lp:
        print("  ✗ ERROR: No se encontró la hoja 'Formatos' en layout_places.")
        return False

    fila_dest = 1
    filas_formatos = 0

    for fi, fila_data in enumerate(rows_formatos_lp, start=1):
        val_a = fila_data[0] if len(fila_data) > 0 else None
        val_b = fila_data[1] if len(fila_data) > 1 else None
        val_c = fila_data[2] if len(fila_data) > 2 else None
        val_d = fila_data[3] if len(fila_data) > 3 else None

        if fi > 1 and all(val is None or str(val).strip() == "" for val in [val_a, val_b, val_c, val_d]):
            continue

        ws_dst.cell(row=fila_dest, column=1).value = val_a
        ws_dst.cell(row=fila_dest, column=2).value = val_b
        ws_dst.cell(row=fila_dest, column=3).value = val_c
        ws_dst.cell(row=fila_dest, column=5).value = val_d  # columna E destino

        fila_dest += 1
        filas_formatos += 1

    print(f"  ✓ Copiadas {filas_formatos} filas de 'Formatos' (A-C -> A-C, D -> E)")

    # Encabezados y fórmulas chain / channel
    ws_dst.cell(row=1, column=4).value = "chain"   # col D
    ws_dst.cell(row=1, column=6).value = "channel" # col F

    # --- Bloque 2: Lugares (layout_places hoja "Lugares" columnas A,B,N,O,P -> G:K) ---
    rows_lug_lp = hojas_lp.get("Lugares", [])
    if not rows_lug_lp:
        print("  ✗ ERROR: No se encontró la hoja 'Lugares' en layout_places.")
        return False

    filas_lug_layout = 0
    fila_dest = 1
    for fi, fila_data in enumerate(rows_lug_lp, start=1):
        cols_necesarias = [0, 1, 13, 14, 15]  # A(0),B(1),N(13),O(14),P(15) (0-based)
        vals = [fila_data[c] if c < len(fila_data) else None for c in cols_necesarias]
        if fi > 1 and all(val is None or str(val).strip() == "" for val in vals):
            continue
        for idx, val in enumerate(vals):
            ws_dst.cell(row=fila_dest, column=7 + idx).value = val  # G=7
        fila_dest += 1
        filas_lug_layout += 1

    print(f"  ✓ Copiadas {filas_lug_layout} filas de 'Lugares' (A,B,N,O,P -> G:K)")

    # --- Bloque 3: LUGARES (matriz) columnas F,G -> L,M ---
    # Copiar columnas F y G completas de LUGARES a L y M de FORMATOS
    # F y G tienen fórmulas XLOOKUP, así que calculamos los valores:
    # F = XLOOKUP(D, C:C, A:A) -> busca D en C, devuelve A
    # G = XLOOKUP(F, A:A, B:B) -> busca F en A, devuelve B
    
    # Construir mapas con múltiples formatos de clave para mayor robustez
    map_c_a = {}  # C -> A (normalizado)
    map_a_b = {}  # A -> B (normalizado)
    
    for fila in range(2, ws_lugares.max_row + 1):
        val_a = ws_lugares.cell(row=fila, column=1).value
        val_b = ws_lugares.cell(row=fila, column=2).value
        val_c = ws_lugares.cell(row=fila, column=3).value
        
        # Guardar en mapa C -> A con múltiples formatos de clave
        if val_c is not None and val_a is not None:
            # Clave como string limpio
            key_c = str(val_c).strip()
            map_c_a[key_c] = val_a
            # También guardar como número si es posible
            try:
                key_c_num = str(int(float(val_c)))
                map_c_a[key_c_num] = val_a
            except (ValueError, TypeError):
                pass
        
        # Guardar en mapa A -> B
        if val_a is not None and val_b is not None:
            key_a = str(val_a).strip()
            map_a_b[key_a] = val_b
            try:
                key_a_num = str(int(float(val_a)))
                map_a_b[key_a_num] = val_b
            except (ValueError, TypeError):
                pass

    print(f"    Mapa C->A construido: {len(map_c_a)} entradas")
    print(f"    Mapa A->B construido: {len(map_a_b)} entradas")

    # Encontrar última fila con datos en LUGARES
    max_fila_lugares = 1
    for fila in range(1, ws_lugares.max_row + 1):
        val_d = ws_lugares.cell(row=fila, column=4).value
        val_f = ws_lugares.cell(row=fila, column=6).value
        val_g = ws_lugares.cell(row=fila, column=7).value
        if val_d is not None or val_f is not None or val_g is not None:
            max_fila_lugares = fila

    # Función para buscar en mapa con múltiples formatos
    def buscar_en_mapa(mapa, valor):
        if valor is None:
            return None
        # Intentar como string limpio
        key = str(valor).strip()
        if key in mapa:
            return mapa[key]
        # Intentar como número entero
        try:
            key_num = str(int(float(valor)))
            if key_num in mapa:
                return mapa[key_num]
        except (ValueError, TypeError):
            pass
        return None

    # Copiar todas las filas de F,G a L,M
    filas_lugares = 0
    filas_calculadas = 0
    for fila in range(1, max_fila_lugares + 1):
        if fila == 1:
            # Fila 1: encabezados
            val_f = ws_lugares.cell(row=1, column=6).value
            val_g = ws_lugares.cell(row=1, column=7).value
            # Si son fórmulas, usar texto fijo
            if isinstance(val_f, str) and val_f.startswith("="):
                val_f = "ID Sto"
            if isinstance(val_g, str) and val_g.startswith("="):
                val_g = "Nombre Sto"
        else:
            # Filas 2+: calcular valores replicando XLOOKUP
            # F = XLOOKUP(D, C:C, A:A)
            # G = XLOOKUP(F, A:A, B:B)
            val_d = ws_lugares.cell(row=fila, column=4).value
            
            # Calcular F buscando D en C para obtener A
            val_f = buscar_en_mapa(map_c_a, val_d)
            
            # Calcular G buscando F en A para obtener B
            val_g = None
            if val_f is not None:
                val_g = buscar_en_mapa(map_a_b, val_f)
                filas_calculadas += 1
        
        ws_dst.cell(row=fila, column=12).value = val_f  # L
        ws_dst.cell(row=fila, column=13).value = val_g  # M
        filas_lugares += 1

    print(f"  ✓ Copiadas {filas_lugares} filas de LUGARES F,G -> FORMATOS L,M ({filas_calculadas} calculadas)")

    # Encabezados N a W
    headers_values = {
        14: "ID Canal",
        15: "Canal",
        16: "ID Cadena",
        17: "Cadena",
        18: "ID Formato",
        19: "Formato",
        20: "LLAVE lugares",
        21: "LLAVE formatos",
        22: "ID Formato final",
        23: "Formato final",
    }
    for col_idx, text in headers_values.items():
        ws_dst.cell(row=1, column=col_idx).value = text

    # --- Aplicar fórmulas ---
    max_fila = max(filas_formatos, filas_lug_layout, filas_lugares)
    if max_fila < 2:
        print("  ! Sin datos para fórmulas en FORMATOS")
        return True

    # Fórmulas principales (hasta el máximo de datos disponibles)
    for fila in range(2, max_fila + 1):
        ws_dst.cell(row=fila, column=4).value = f"=_xlfn.XLOOKUP(C{fila},'CADENA CANAL FORMATO'!C:C,'CADENA CANAL FORMATO'!D:D)"  # chain
        ws_dst.cell(row=fila, column=6).value = f"=_xlfn.XLOOKUP(E{fila},'CADENA CANAL FORMATO'!E:E,'CADENA CANAL FORMATO'!F:F)"  # channel

        # N-S
        ws_dst.cell(row=fila, column=14).value = f"=_xlfn.XLOOKUP(O{fila},F:F,E:E)"          # N ID Canal
        ws_dst.cell(row=fila, column=15).value = f"=_xlfn.XLOOKUP(L{fila},$G:$G,$I:$I)"      # O Canal
        ws_dst.cell(row=fila, column=16).value = f"=_xlfn.XLOOKUP(Q{fila},D:D,C:C)"          # P ID Cadena
        ws_dst.cell(row=fila, column=17).value = f"=_xlfn.XLOOKUP(L{fila},$G:$G,$J:$J)"      # Q Cadena
        ws_dst.cell(row=fila, column=18).value = f"=_xlfn.XLOOKUP(S{fila},B:B,A:A)"          # R ID Formato
        ws_dst.cell(row=fila, column=19).value = f"=_xlfn.XLOOKUP(L{fila},G:G,K:K)"          # S Formato

        # T-U
        ws_dst.cell(row=fila, column=20).value = f"=S{fila}&\" / \"&P{fila}&\" / \"&N{fila}"
        ws_dst.cell(row=fila, column=21).value = f"=B{fila}&\" / \"&C{fila}&\" / \"&E{fila}"

    # Fórmulas V-W solo hasta donde hay datos en L/M (filas_lugares)
    max_fila_lm = max(filas_lugares, 1)
    for fila in range(2, max_fila_lm + 1):
        ws_dst.cell(row=fila, column=22).value = f"=_xlfn.XLOOKUP(T{fila},U:U,A:A)"
        ws_dst.cell(row=fila, column=23).value = f"=_xlfn.XLOOKUP(V{fila},A:A,B:B)"

    # --- Encabezados y colores ---
    headers = {
        1: ("A", "C", "C6EFCE"),   # A-C
        4: ("D", "F", "BDD7EE"),   # D-F
        7: ("G", "K", "FFF2CC"),   # G-K
        12: ("L", "M", "F8CBAD"),  # L-M
        14: ("N", "S", "F4B084"),  # N-S
        20: ("T", "U", "D9D2E9"),  # T-U
        22: ("V", "W", "D9D9D9"),  # V-W
    }

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center")
    font_header = Font(bold=True, color="000000")

    for start_col, (col_ini, col_fin, color_hex) in headers.items():
        fill = PatternFill(fill_type="solid", fgColor=color_hex)
        for col in range(start_col, column_index_from_string(col_fin) + 1):
            celda = ws_dst.cell(row=1, column=col)
            celda.fill = fill
            celda.font = font_header
            celda.alignment = align
            celda.border = border

    print("  ✓ FORMATOS completado con fórmulas y encabezados coloreados")
    return True


def procesar_configuracion_anaquel(wb, ruta_matriz):
    """
    Procesa la hoja CONFIGURACIÓN DE ANAQUEL:
    1. Inserta 4 columnas después de E (recorre el contenido existente)
    2. Copia FORMATOS V,W -> F,G (encabezados en fila 4, datos debajo)
    3. Copia LUGARES F,G -> H,I (encabezados en fila 4, datos debajo)
    4. Copia PRODUCTOS filas 4,5,6 desde columna E -> CONFIGURACIÓN DE ANAQUEL filas 2,3,4 desde columna K
    """
    print("\n--- Procesando CONFIGURACIÓN DE ANAQUEL ---")
    
    # Verificar que existan las hojas necesarias
    hojas_requeridas = ["CONFIGURACIÓN DE ANAQUEL", "FORMATOS", "LUGARES", "PRODUCTOS", "CADENA CANAL FORMATO"]
    for hoja in hojas_requeridas:
        if hoja not in wb.sheetnames:
            print(f"  ✗ ERROR: No se encontró la hoja '{hoja}'")
            return False
    
    ws_config = wb["CONFIGURACIÓN DE ANAQUEL"]
    ws_formatos = wb["FORMATOS"]
    ws_lugares = wb["LUGARES"]
    ws_productos = wb["PRODUCTOS"]
    ws_ccf = wb["CADENA CANAL FORMATO"]

    # Helper para copiar estilo además del valor (sin tocar fórmulas; se pega valor y formato)
    from copy import copy
    def copiar_valor_y_estilo(celda_val, celda_style_src, celda_dst):
        celda_dst.value = celda_val
        try:
            celda_dst._style = copy(celda_style_src._style)
        except Exception:
            pass

    # Limpiar fila 2 (contenido + formato) antes de pegar, sin borrar filas
    blank_style = copy(openpyxl.Workbook().active.cell(1, 1)._style)
    if ws_config.merged_cells.ranges:
        for rng in list(ws_config.merged_cells.ranges):
            if rng.min_row <= 2 <= rng.max_row:
                try:
                    ws_config.unmerge_cells(str(rng))
                except Exception:
                    pass
    for col_idx in range(1, ws_config.max_column + 1):
        celda = ws_config.cell(row=2, column=col_idx)
        celda.value = None
        celda._style = copy(blank_style)

    # 1. Insertar 4 columnas después de E (columna 5), es decir, insertar en columna 6 (F)
    print("  Insertando 4 columnas después de E...")
    ws_config.insert_cols(6, 4)  # Inserta 4 columnas a partir de la columna 6 (F)
    print("  ✓ 4 columnas insertadas (el contenido existente se recorrió)")

    # 2. Eliminar columnas Z a AK (26 a 37) después de la inserción previa
    print("  Eliminando columnas Z a AK en CONFIGURACIÓN DE ANAQUEL...")
    ws_config.delete_cols(26, 12)  # Z (26) a AK (37) inclusive
    print("  ✓ Columnas Z a AK eliminadas")

    # 3. Insertar 1 columna después de la columna CH vigente (se calcula tras el borrado)
    print("  Insertando 1 columna después de CH...")
    col_ch = column_index_from_string("CH")
    ws_config.insert_cols(col_ch + 1, 1)
    print("  ✓ 1 columna insertada después de CH (contenido recorrido)")
    
    # ========== CONSTRUIR MAPAS PARA CALCULAR VALORES ==========
    # Estos mapas replican la lógica de los XLOOKUP en Python
    
    # --- Mapas desde LUGARES (para calcular F y G de LUGARES) ---
    # F = XLOOKUP(D, C:C, A:A) -> busca D en C, devuelve A
    # G = XLOOKUP(F, A:A, B:B) -> busca F en A, devuelve B
    map_lugares_c_a = {}  # C -> A (ID lugar)
    map_lugares_a_b = {}  # A -> B (Nombre lugar)
    for fila in range(2, ws_lugares.max_row + 1):
        val_a = ws_lugares.cell(row=fila, column=1).value
        val_b = ws_lugares.cell(row=fila, column=2).value
        val_c = ws_lugares.cell(row=fila, column=3).value
        if val_c not in (None, "") and val_a not in (None, ""):
            map_lugares_c_a[str(val_c).strip()] = val_a
        if val_a not in (None, "") and val_b not in (None, ""):
            map_lugares_a_b[str(val_a).strip()] = val_b
    
    # --- Mapas desde CADENA CANAL FORMATO ---
    map_ccf_c_d = {}  # C (ID Cadena) -> D (Nombre Cadena)
    map_ccf_e_f = {}  # E (ID Canal) -> F (Nombre Canal)
    map_ccf_a_b = {}  # A (ID Formato) -> B (Nombre Formato)
    map_ccf_b_a = {}  # B (Nombre Formato) -> A (ID Formato)
    for fila in range(2, ws_ccf.max_row + 1):
        val_a = ws_ccf.cell(row=fila, column=1).value  # ID Formato
        val_b = ws_ccf.cell(row=fila, column=2).value  # Nombre Formato
        val_c = ws_ccf.cell(row=fila, column=3).value  # ID Cadena
        val_d = ws_ccf.cell(row=fila, column=4).value  # Nombre Cadena
        val_e = ws_ccf.cell(row=fila, column=5).value  # ID Canal
        val_f = ws_ccf.cell(row=fila, column=6).value  # Nombre Canal
        
        if val_a not in (None, "") and val_b not in (None, ""):
            map_ccf_a_b[str(val_a).strip()] = val_b
            map_ccf_b_a[str(val_b).strip()] = val_a
        if val_c not in (None, "") and val_d not in (None, ""):
            map_ccf_c_d[str(val_c).strip()] = val_d
        if val_e not in (None, "") and val_f not in (None, ""):
            map_ccf_e_f[str(val_e).strip()] = val_f
    
    # --- Mapas desde FORMATOS (columnas base sin fórmulas) ---
    # A = ID Formato, B = Nombre Formato, C = ID Cadena, E = ID Canal
    # G = ID lugar layout, I = Canal, J = Cadena, K = Formato (del layout)
    # L = ID Sto (calculado desde LUGARES)
    map_formatos_u_a = {}  # U (llave) -> A (ID Formato)
    # U = B & " / " & C & " / " & E
    for fila in range(2, ws_formatos.max_row + 1):
        val_a = ws_formatos.cell(row=fila, column=1).value  # ID Formato
        val_b = ws_formatos.cell(row=fila, column=2).value  # Nombre Formato
        val_c = ws_formatos.cell(row=fila, column=3).value  # ID Cadena
        val_e = ws_formatos.cell(row=fila, column=5).value  # ID Canal
        if all(v not in (None, "") for v in [val_a, val_b, val_c, val_e]):
            llave_u = f"{val_b} / {val_c} / {val_e}"
            map_formatos_u_a[llave_u] = val_a
    
    # Mapa G -> (I, J, K) desde FORMATOS para calcular T
    map_formatos_g_ijk = {}  # ID lugar -> (Canal, Cadena, Formato)
    for fila in range(2, ws_formatos.max_row + 1):
        val_g = ws_formatos.cell(row=fila, column=7).value   # G = ID lugar
        val_i = ws_formatos.cell(row=fila, column=9).value   # I = Canal
        val_j = ws_formatos.cell(row=fila, column=10).value  # J = Cadena
        val_k = ws_formatos.cell(row=fila, column=11).value  # K = Formato
        if val_g not in (None, ""):
            map_formatos_g_ijk[str(val_g).strip()] = (val_i, val_j, val_k)
    
    # Mapas para convertir nombres a IDs
    map_canal_nombre_a_id = {}
    map_cadena_nombre_a_id = {}
    for fila in range(2, ws_ccf.max_row + 1):
        val_c = ws_ccf.cell(row=fila, column=3).value  # ID Cadena
        val_d = ws_ccf.cell(row=fila, column=4).value  # Nombre Cadena
        val_e = ws_ccf.cell(row=fila, column=5).value  # ID Canal
        val_f = ws_ccf.cell(row=fila, column=6).value  # Nombre Canal
        if val_d not in (None, "") and val_c not in (None, ""):
            map_cadena_nombre_a_id[str(val_d).strip()] = val_c
        if val_f not in (None, "") and val_e not in (None, ""):
            map_canal_nombre_a_id[str(val_f).strip()] = val_e
    
    # ========== COPIAR FORMATOS V,W -> F,G ==========
    # V = XLOOKUP(T, U:U, A:A) donde T = S & " / " & P & " / " & N
    # W = XLOOKUP(V, A:A, B:B)
    print("  Copiando FORMATOS V,W -> F,G (calculando valores)...")
    
    # Encontrar última fila con datos en FORMATOS columna L (ID Sto)
    max_fila_formatos = 1
    for fila in range(1, ws_formatos.max_row + 1):
        val_l = ws_formatos.cell(row=fila, column=12).value
        if val_l is not None:
            max_fila_formatos = fila
    
    # Copiar encabezados
    copiar_valor_y_estilo("ID Formato final", ws_formatos.cell(row=1, column=22), ws_config.cell(row=4, column=6))
    copiar_valor_y_estilo("Formato final", ws_formatos.cell(row=1, column=23), ws_config.cell(row=4, column=7))
    
    filas_copiadas_formatos = 0
    for fila_src in range(2, max_fila_formatos + 1):
        fila_dest = fila_src + 3
        
        # Obtener L (ID Sto) para buscar el lugar en el mapa
        val_l = ws_formatos.cell(row=fila_src, column=12).value
        
        val_v = None
        val_w = None
        
        if val_l not in (None, ""):
            # Buscar datos del lugar desde el mapa G -> (I, J, K)
            datos_lugar = map_formatos_g_ijk.get(str(val_l).strip())
            if datos_lugar:
                canal_nombre, cadena_nombre, formato_nombre = datos_lugar
                
                # Convertir nombres a IDs para calcular T
                id_canal = map_canal_nombre_a_id.get(str(canal_nombre).strip()) if canal_nombre else None
                id_cadena = map_cadena_nombre_a_id.get(str(cadena_nombre).strip()) if cadena_nombre else None
                
                # T = Formato & " / " & ID Cadena & " / " & ID Canal
                if formato_nombre and id_cadena and id_canal:
                    llave_t = f"{formato_nombre} / {id_cadena} / {id_canal}"
                    # V = XLOOKUP(T, U:U, A:A)
                    val_v = map_formatos_u_a.get(llave_t)
                    if val_v is not None:
                        # W = XLOOKUP(V, A:A, B:B)
                        val_w = map_ccf_a_b.get(str(val_v).strip())
        
        copiar_valor_y_estilo(val_v, ws_formatos.cell(row=fila_src, column=22), ws_config.cell(row=fila_dest, column=6))
        copiar_valor_y_estilo(val_w, ws_formatos.cell(row=fila_src, column=23), ws_config.cell(row=fila_dest, column=7))
        filas_copiadas_formatos += 1
    
    print(f"  ✓ Copiadas {filas_copiadas_formatos} filas de FORMATOS V,W -> F,G")
    
    # ========== COPIAR LUGARES F,G -> H,I ==========
    # F = XLOOKUP(D, C:C, A:A)
    # G = XLOOKUP(F, A:A, B:B)
    print("  Copiando LUGARES F,G -> H,I (calculando valores)...")
    
    # Encontrar última fila con datos en LUGARES columna D
    max_fila_lugares = 1
    for fila in range(1, ws_lugares.max_row + 1):
        val_d = ws_lugares.cell(row=fila, column=4).value
        if val_d is not None:
            max_fila_lugares = fila
    
    # Copiar encabezados (fila 1 de LUGARES -> fila 4 de CONFIG)
    copiar_valor_y_estilo("ID Sto", ws_lugares.cell(row=1, column=6), ws_config.cell(row=4, column=8))  # F -> H
    copiar_valor_y_estilo("Nombre Sto", ws_lugares.cell(row=1, column=7), ws_config.cell(row=4, column=9))  # G -> I
    
    # Copiar datos calculados (fila 2+ de LUGARES -> fila 5+ de CONFIG)
    filas_copiadas_lugares = 0
    for fila_src in range(2, max_fila_lugares + 1):
        fila_dest = fila_src + 3  # fila 2 -> fila 5, fila 3 -> fila 6, etc.
        
        # Calcular F y G usando los mapas (replicando XLOOKUP)
        # F = XLOOKUP(D, C:C, A:A)
        # G = XLOOKUP(F, A:A, B:B)
        val_d = ws_lugares.cell(row=fila_src, column=4).value
        val_f = None
        val_g = None
        
        if val_d not in (None, ""):
            val_f = map_lugares_c_a.get(str(val_d).strip())
            if val_f is not None:
                val_g = map_lugares_a_b.get(str(val_f).strip())
        
        copiar_valor_y_estilo(val_f, ws_lugares.cell(row=fila_src, column=6), ws_config.cell(row=fila_dest, column=8))  # H
        copiar_valor_y_estilo(val_g, ws_lugares.cell(row=fila_src, column=7), ws_config.cell(row=fila_dest, column=9))  # I
        filas_copiadas_lugares += 1
    
    print(f"  ✓ Copiadas {filas_copiadas_lugares} filas de LUGARES F,G -> H,I")
    
    # ========== COPIAR PRODUCTOS filas 4,5,6 -> CONFIG filas 2,3,4 desde K ==========
    # E = col 5 en PRODUCTOS
    # K = col 11 en CONFIGURACIÓN DE ANAQUEL
    # Las filas 5 y 6 tienen fórmulas XLOOKUP que necesitamos calcular:
    # Fila 5 (ID): =XLOOKUP(col2, $C$8:$C$1000, $A$8:$A$1000)
    # Fila 6 (Nombre): =XLOOKUP(col2, $C$8:$C$1000, $B$8:$B$1000)
    print("  Copiando PRODUCTOS filas 4,5,6 desde E -> CONFIG filas 2,3,4 desde K...")
    
    # Construir mapa C -> (A, B) desde PRODUCTOS filas 8-1000
    # C = SKU, A = ID, B = Nombre
    map_productos_c_ab = {}  # C (SKU) -> (A=ID, B=Nombre)
    for fila in range(8, min(ws_productos.max_row + 1, 1001)):
        val_a = ws_productos.cell(row=fila, column=1).value  # A = ID
        val_b = ws_productos.cell(row=fila, column=2).value  # B = Nombre
        val_c = ws_productos.cell(row=fila, column=3).value  # C = SKU
        if val_c not in (None, ""):
            map_productos_c_ab[str(val_c).strip()] = (val_a, val_b)
    
    print(f"    Mapa de productos construido: {len(map_productos_c_ab)} entradas")
    
    # Encontrar última columna con datos en PRODUCTOS fila 4 desde columna E
    max_col_productos = 5  # Mínimo columna E
    for col in range(5, ws_productos.max_column + 1):
        # Verificar si hay datos en filas 2, 4, 5 o 6
        val_2 = ws_productos.cell(row=2, column=col).value
        val_4 = ws_productos.cell(row=4, column=col).value
        val_5 = ws_productos.cell(row=5, column=col).value
        val_6 = ws_productos.cell(row=6, column=col).value
        if val_2 is not None or val_4 is not None or val_5 is not None or val_6 is not None:
            max_col_productos = col
    
    # Deshacer celdas combinadas en el área de destino (filas 2-4, desde columna K)
    from openpyxl.cell.cell import MergedCell

    col_dest_min = 11
    col_dest_max = 11 + (max_col_productos - 5)

    def _unmerge_if_merged(row_idx, col_idx):
        cell = ws_config.cell(row=row_idx, column=col_idx)
        if not isinstance(cell, MergedCell):
            return False
        for merged_range in list(ws_config.merged_cells.ranges):
            if (merged_range.min_row <= row_idx <= merged_range.max_row and
                merged_range.min_col <= col_idx <= merged_range.max_col):
                ws_config.unmerge_cells(str(merged_range))
                return True
        return False

    merges_removed = 0
    for r in (2, 3, 4):
        for c in range(col_dest_min, col_dest_max + 1):
            try:
                if _unmerge_if_merged(r, c):
                    merges_removed += 1
            except Exception:
                pass

    if merges_removed:
        print(f"  ✓ Deshecho {merges_removed} merges que bloqueaban escritura")
    
    # Función para obtener valor (calculando si es fórmula XLOOKUP)
    def obtener_valor_calculado(celda, fila_origen, col_idx):
        val = celda.value
        
        # Si no es fórmula, devolver el valor directo
        if not isinstance(val, str) or not val.startswith("="):
            return val
        
        # Si es fórmula en fila 5 o 6, calcular usando el mapa
        if fila_origen in (5, 6):
            # Obtener el valor de fila 2 en la misma columna (es el SKU a buscar)
            val_fila2 = ws_productos.cell(row=2, column=col_idx).value
            if val_fila2 not in (None, ""):
                datos = map_productos_c_ab.get(str(val_fila2).strip())
                if datos:
                    if fila_origen == 5:  # Fila 5 = ID (columna A)
                        return datos[0]
                    else:  # Fila 6 = Nombre (columna B)
                        return datos[1]
        
        # Si es otra fórmula o no se encontró, devolver None
        return None
    
    cols_copiadas = 0
    for col_src in range(5, max_col_productos + 1):  # Desde E (col 5)
        col_dest = 11 + (col_src - 5)  # K (col 11) + offset

        # Asegurar que las celdas destino no sean MergedCell
        for r in (2, 3, 4):
            try:
                _unmerge_if_merged(r, col_dest)
            except Exception:
                pass
        
        # Fila 4 PRODUCTOS -> Fila 2 CONFIG
        val_4 = obtener_valor_calculado(ws_productos.cell(row=4, column=col_src), 4, col_src)
        copiar_valor_y_estilo(val_4, ws_productos.cell(row=4, column=col_src), ws_config.cell(row=2, column=col_dest))
        
        # Fila 5 PRODUCTOS -> Fila 3 CONFIG (ID calculado)
        val_5 = obtener_valor_calculado(ws_productos.cell(row=5, column=col_src), 5, col_src)
        copiar_valor_y_estilo(val_5, ws_productos.cell(row=5, column=col_src), ws_config.cell(row=3, column=col_dest))
        
        # Fila 6 PRODUCTOS -> Fila 4 CONFIG (Nombre calculado)
        val_6 = obtener_valor_calculado(ws_productos.cell(row=6, column=col_src), 6, col_src)
        copiar_valor_y_estilo(val_6, ws_productos.cell(row=6, column=col_src), ws_config.cell(row=4, column=col_dest))
        
        cols_copiadas += 1
    
    print(f"  ✓ Copiadas {cols_copiadas} columnas de PRODUCTOS (filas 4-6 desde E) -> CONFIG (filas 2-4 desde K)")
    
    print("  ✓ CONFIGURACIÓN DE ANAQUEL procesado completamente")
    return True


def procesar_matriz(ruta_matriz, ruta_layout_products, ruta_layout_places, ruta_base):
    """Procesa el archivo MATRIZ DE CATALOGACIÓN."""
    print("\n" + "="*50)
    print("Procesando MATRIZ DE CATALOGACIÓN...")
    print("="*50)
    
    # Cargar el libro de Excel
    print(f"\nAbriendo archivo: {os.path.basename(ruta_matriz)}")
    wb = openpyxl.load_workbook(ruta_matriz)
    
    print(f"Hojas existentes: {wb.sheetnames}")
    
    # 1. Copiar hoja CONFIGURACIÓN DE ANAQUEL -> CONFIGURACIÓN DE ANAQUEL 1
    print("\n--- Copiando hoja de respaldo ---")
    if not copiar_hoja(wb, "CONFIGURACIÓN DE ANAQUEL", "CONFIGURACIÓN DE ANAQUEL 1"):
        wb.close()
        return False
    
    # 2. Crear / asegurar y limpiar hojas de trabajo
    print("\n--- Creando y limpiando hojas ---")
    hojas_nuevas = ["PRODUCTOS", "LUGARES", "FORMATOS", "CADENA CANAL FORMATO"]

    for nombre_hoja in hojas_nuevas:
        hoja = crear_hoja(wb, nombre_hoja)
        limpiar_hoja(hoja)

    # 3. Copiar productos desde layout_products
    ok_productos, filas_copiadas = copiar_productos(ruta_layout_products, wb)
    if not ok_productos:
        wb.close()
        return False

    # 4. Copiar configuración de anaquel auxiliar hacia PRODUCTOS
    ok_config, ancho_copiado, col_inicio_copiado = copiar_configuracion_anaquel(wb)
    if not ok_config:
        wb.close()
        return False
    
    # 5. Eliminar columnas de productos discontinuados (300ml) y de competencia (Total xxx)
    ws_productos = wb["PRODUCTOS"]
    eliminar_columnas_dinamico(ws_productos, col_inicio_copiado)
    
    # 6. Llenar espacios vacíos con datos de Productos de Sams
    llenar_espacios_vacios_productos_sams(ws_productos, col_inicio_copiado, ruta_base)

    # 6b. Asegurar color azul en fila 5 desde la columna E en adelante
    ancho_total = ws_productos.max_column - col_inicio_copiado + 1
    try:
        aplicar_color_fila5(ws_productos, col_inicio_copiado, ancho_total, color_hex="00B0F0")
    except Exception:
        pass
    
    # 7. Establecer fórmulas solo donde no haya valores
    establecer_formulas_id_nombre(ws_productos, col_inicio_copiado)
    
    # 8. Pre-cargar layout_places UNA sola vez en memoria para las tres funciones
    hojas_lp = {}
    if ruta_layout_places and os.path.exists(ruta_layout_places):
        print("\n--- Pre-cargando layout_places en memoria ---")
        hojas_lp = _leer_wb_hojas(ruta_layout_places, "Lugares", "Formatos", "Cadenas", "Canales")
        print(f"  ✓ Hojas cargadas: {list(hojas_lp.keys())}")
    
    # 9. Procesar hoja LUGARES
    if ruta_layout_places:
        procesar_lugares(wb, ruta_layout_places, hojas_lp=hojas_lp)
    
    # 10. Procesar hoja CADENA CANAL FORMATO
    if ruta_layout_places:
        procesar_cadena_canal_formato(wb, ruta_layout_places, hojas_lp=hojas_lp)

    # 11. Procesar hoja FORMATOS
    if ruta_layout_places:
        procesar_formatos(wb, ruta_layout_places, hojas_lp=hojas_lp)
    
    # 12. Guardar archivo intermedio para que las fórmulas estén disponibles
    print("\n--- Guardando archivo intermedio ---")
    wb.save(ruta_matriz)
    wb.close()
    print("  ✓ Archivo guardado (fórmulas escritas)")
    
    # 13. Reabrir el archivo para procesar CONFIGURACIÓN DE ANAQUEL
    print("\n--- Reabriendo archivo para procesar CONFIG ---")
    wb = openpyxl.load_workbook(ruta_matriz)
    
    # 14. Procesar hoja CONFIGURACIÓN DE ANAQUEL
    procesar_configuracion_anaquel(wb, ruta_matriz)

    # 14b. Rellenar columnas CJ-CT con 1s antes del filtro
    rellenar_unos_configuracion_anaquel(wb)

    # 15. Aplicar filtro de formato (ID 4770)
    procesar_filtro_formato(wb)

    # 16. Marcar producto general
    procesar_producto_general(wb)
    
    # Guardar el archivo
    print("\n--- Guardando cambios ---")
    wb.save(ruta_matriz)
    print(f"✓ Archivo guardado: {os.path.basename(ruta_matriz)}")
    
    wb.close()
    
    print("\n" + "="*50)
    print("RESUMEN")
    print("="*50)
    print("✓ Hoja de respaldo creada/mantenida: CONFIGURACIÓN DE ANAQUEL 1 (contenido preservado)")
    print("✓ Hojas creadas/limpiadas: PRODUCTOS, LUGARES, FORMATOS, CADENA CANAL FORMATO")
    print("✓ Productos copiados: {filas_copiadas} filas (cols A,B,E -> A,B,C desde fila 8)")
    print("✓ Columna C normalizada (texto a número si aplica) en PRODUCTOS")
    print("✓ Copiado CONFIGURACIÓN DE ANAQUEL 1 -> PRODUCTOS (G2.. hasta última col con datos, filas 2-4, con estilos)")
    print("✓ Fórmulas BUSCARX colocadas en PRODUCTOS fila 5, arrastradas a la derecha")
    print("✓ LUGARES procesado: layout_places cols A,B,F -> A,B,C con fórmulas BUSCARX en F,G")
    print("✓ CADENA CANAL FORMATO procesado: hojas Formatos, Cadenas, Canales copiadas")
    print("✓ FORMATOS procesado: datos y fórmulas completados")
    print("✓ CONFIGURACIÓN DE ANAQUEL procesado: 4 cols insertadas, datos de FORMATOS/LUGARES/PRODUCTOS copiados")
    print("✓ Producto general marcado (incluye duplicado CH->CI)")
    print("✓ Formato 4770 ajustado (CU-DG=1, CD=1, CJ-CT limpio)")
    
    return True


