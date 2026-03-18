"""
Script para actualizar columnas específicas para todas las filas del formato 4770.

Este script filtra las filas por ID Formato Storecheck (columna F) y:
1. Coloca 1 en todas las columnas SAMS
2. Elimina el contenido de las columnas de COMPETENCIA
"""

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

# Columna desde la que comienzan los grupos de productos en CONFIG DE ANAQUEL
_COL_INICIO_PRODUCTOS_CONFIG = column_index_from_string("K")


def _detectar_secciones(ws):
    """Detecta dinámicamente los rangos de competencia y SAMS en CONFIG DE ANAQUEL.

    Escanea la fila 2 (nombre de grupo) desde la columna K en busca de los marcadores
    que escribe `llenar_espacios_vacios_productos_sams`:
      · 'Competencia'    → primer marcador de columnas de competencia
      · 'Productos Sams' → primer marcador de columnas SAMS (o cualquier texto con 'sams')

    Retorna:
        (comp_inicio, comp_fin, sams_inicio, sams_fin)
    Cualquier par puede ser (None, None) si la sección no existe.
    """
    comp_inicio = comp_fin = None
    sams_inicio = sams_fin = None

    for col in range(_COL_INICIO_PRODUCTOS_CONFIG, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val is None:
            continue
        s = str(val).strip().lower()
        if "competencia" in s and comp_inicio is None:
            comp_inicio = col
        elif ("sams" in s) and sams_inicio is None:
            sams_inicio = col

    # Extensión de competencia: hasta la columna previa a SAMS (o hasta el último dato)
    if comp_inicio is not None:
        if sams_inicio is not None:
            comp_fin = sams_inicio - 1
        else:
            comp_fin = comp_inicio
            for col in range(comp_inicio, ws.max_column + 1):
                if (ws.cell(row=3, column=col).value is not None or
                        ws.cell(row=4, column=col).value is not None):
                    comp_fin = col

    # Extensión de SAMS: hasta el último col con dato en filas 3 o 4
    if sams_inicio is not None:
        sams_fin = sams_inicio
        for col in range(sams_inicio, ws.max_column + 1):
            if (ws.cell(row=3, column=col).value is not None or
                    ws.cell(row=4, column=col).value is not None):
                sams_fin = col

    return comp_inicio, comp_fin, sams_inicio, sams_fin


def rellenar_unos_configuracion_anaquel(wb):
    """Escribe 1s en las columnas de COMPETENCIA detectadas dinámicamente, comenzando desde fila 5."""
    if "CONFIGURACIÓN DE ANAQUEL" not in wb.sheetnames:
        print("  ! No se encontró CONFIGURACIÓN DE ANAQUEL para llenar 1s")
        return

    ws = wb["CONFIGURACIÓN DE ANAQUEL"]

    comp_inicio, comp_fin, _, _ = _detectar_secciones(ws)

    if comp_inicio is None:
        print("  ! No se detectaron columnas de competencia en CONFIG DE ANAQUEL")
        return

    filas = 0
    for fila in range(5, ws.max_row + 1):
        for col in range(comp_inicio, comp_fin + 1):
            ws.cell(row=fila, column=col).value = 1
        filas += 1

    print(f"  ✓ Columnas competencia "
          f"{get_column_letter(comp_inicio)}-{get_column_letter(comp_fin)} "
          f"llenas de 1s en filas 5-{ws.max_row} ({filas} filas afectadas)")


def procesar_filtro_formato(wb):
    """Para ID formato 4770 coloca 1 en SAMS y limpia COMPETENCIA."""
    print("\n--- Procesando filtro de formato (ID 4770) ---")

    if "CONFIGURACIÓN DE ANAQUEL" not in wb.sheetnames:
        print("  ✗ ERROR: No se encontró la hoja 'CONFIGURACIÓN DE ANAQUEL'")
        return False

    ws = wb["CONFIGURACIÓN DE ANAQUEL"]

    comp_inicio, comp_fin, sams_inicio, sams_fin = _detectar_secciones(ws)

    if sams_inicio is None:
        print("  ! No se detectaron columnas SAMS — sin cambios para formato 4770")
        return False

    if comp_inicio is None:
        print("  ! No se detectaron columnas de competencia — solo se llenarán columnas SAMS")

    print(f"  Competencia: {get_column_letter(comp_inicio) if comp_inicio else 'N/A'}-"
          f"{get_column_letter(comp_fin) if comp_fin else 'N/A'}")
    print(f"  SAMS:        {get_column_letter(sams_inicio)}-{get_column_letter(sams_fin)}")

    ID_FORMATO_BUSCAR = 4770
    FILA_INICIO_DATOS = 5

    idx_columna_formato = column_index_from_string('F')

    ultima_fila = ws.max_row
    filas_procesadas = 0
    celdas_actualizadas = 0
    celdas_eliminadas = 0

    for fila in range(FILA_INICIO_DATOS, ultima_fila + 1):
        id_formato = ws.cell(row=fila, column=idx_columna_formato).value
        if id_formato == ID_FORMATO_BUSCAR or str(id_formato).strip() == str(ID_FORMATO_BUSCAR):
            filas_procesadas += 1

            # Llenar columnas SAMS con 1
            for col in range(sams_inicio, sams_fin + 1):
                ws.cell(row=fila, column=col).value = 1
                celdas_actualizadas += 1

            # Limpiar columnas COMPETENCIA
            if comp_inicio is not None:
                for col in range(comp_inicio, comp_fin + 1):
                    ws.cell(row=fila, column=col).value = None
                    celdas_eliminadas += 1

    print(f"  ✓ Filas procesadas: {filas_procesadas}")
    print(f"  ✓ Celdas con 1 colocadas: {celdas_actualizadas}")
    print(f"  ✓ Celdas limpiadas: {celdas_eliminadas}")
    return True


