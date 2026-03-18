"""
Script para actualizar la columna del producto general en la hoja CONFIGURACIÓN DE ANAQUEL.

Este script revisa fila por fila (lugar por lugar) y si hay al menos una celda con valor 1
en el rango de sabores de un grupo, coloca un 1 en la/las columna(s) del
producto general correspondiente.

Los grupos y sus columnas se detectan DINÁMICAMENTE usando marcadores en la fila 2:
- Cada grupo inicia en su marcador de fila 2
- El/los producto(s) general(es) se ubican al final del grupo: columna(s)
    inmediatamente anterior(es) al siguiente marcador
- Regla especial: para "Electrolife Zero Polvo" son 2 columnas generales antes de "Competencia"

No se necesitan rangos hardcodeados: el script se adapta automáticamente si cambian las columnas.
"""

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

NOMBRE_HOJA = "CONFIGURACIÓN DE ANAQUEL"
COL_INICIO_PRODUCTOS = 'K'   # Primera columna donde empiezan los grupos de productos
FILA_INICIO_DATOS = 5        # Fila donde empiezan los datos de lugares

MARCADORES_ORDEN = [
    "Electrolit 625ml",
    "Electrolit 355ml",
    "Electrolit 1000ml",
    "Electrolife zero 625ml",
    "Electrolit Ped 300ml",
    "Electrolit Ped 500ml",
    "Electrolit Six Pack",
    "Electrolife Zero Polvo",
    "Competencia",
]


def _normalizar_texto(texto):
    return " ".join(str(texto).strip().lower().split())



def detectar_grupos(ws):
    """Detecta grupos usando marcadores de fila 2 y calcula generales al final de cada grupo.

    Regla general: 1 columna general al final del rango (antes del siguiente marcador).
    Regla especial: en "Electrolife Zero Polvo" son 2 columnas generales antes de "Competencia".

    Retorna lista de (nombre_grupo, [cols_sabor], [cols_general]).
    """
    col_inicio = column_index_from_string(COL_INICIO_PRODUCTOS)
    max_col = ws.max_column

    marcadores_norm = [_normalizar_texto(m) for m in MARCADORES_ORDEN]
    col_por_marcador = {}

    for col in range(col_inicio, max_col + 1):
        val = ws.cell(row=2, column=col).value
        if val is None or str(val).strip() == "":
            continue
        nombre = str(val).strip()
        nombre_norm = _normalizar_texto(nombre)
        if nombre_norm in marcadores_norm and nombre_norm not in col_por_marcador:
            col_por_marcador[nombre_norm] = col

    grupos = []

    # Procesamos solo grupos de producto (hasta antes de Competencia).
    # Competencia es el marcador terminal del bloque de productos para este cálculo.
    idx_competencia = marcadores_norm.index(_normalizar_texto("Competencia"))

    for i in range(idx_competencia):
        marcador_actual = marcadores_norm[i]
        marcador_siguiente = marcadores_norm[i + 1]

        if marcador_actual not in col_por_marcador or marcador_siguiente not in col_por_marcador:
            continue

        col_inicio_grupo = col_por_marcador[marcador_actual]
        col_siguiente = col_por_marcador[marcador_siguiente]

        cantidad_generales = 2 if marcador_actual == _normalizar_texto("Electrolife Zero Polvo") else 1

        col_general_inicio = col_siguiente - cantidad_generales
        col_general_fin = col_siguiente - 1
        col_sabor_fin = col_general_inicio - 1

        if col_sabor_fin < col_inicio_grupo:
            continue

        cols_sabor = list(range(col_inicio_grupo, col_sabor_fin + 1))
        cols_general = list(range(col_general_inicio, col_general_fin + 1))

        nombre_grupo = MARCADORES_ORDEN[i]
        grupos.append((nombre_grupo, cols_sabor, cols_general))

    return grupos


def _es_uno(valor):
    """True cuando el valor representa un 1 (int/float/str)."""
    if valor is None:
        return False
    if isinstance(valor, (int, float)):
        return valor == 1
    return str(valor).strip() == "1"


def procesar_producto_general(wb):
    """Marca el producto general en CONFIGURACIÓN DE ANAQUEL detectando grupos dinámicamente."""
    print("\n--- Procesando producto general (detección dinámica) ---")

    if NOMBRE_HOJA not in wb.sheetnames:
        print(f"  ✗ ERROR: No se encontró la hoja '{NOMBRE_HOJA}'")
        return False

    ws = wb[NOMBRE_HOJA]
    col_F = column_index_from_string('F')

    grupos = detectar_grupos(ws)
    if not grupos:
        print("  ! No se detectaron grupos de productos.")
        return True

    print(f"  Grupos detectados ({len(grupos)}):")
    for g_name, cols_sabor, cols_general in grupos:
        print(f"  Grupo '{g_name}': {len(cols_sabor)} sabores "
              f"({get_column_letter(cols_sabor[0])}-{get_column_letter(cols_sabor[-1])}), "
              f"{len(cols_general)} general(es) "
              f"({', '.join(get_column_letter(c) for c in cols_general)})")

    ultima_fila = ws.max_row
    total_actualizaciones = 0

    for fila in range(FILA_INICIO_DATOS, ultima_fila + 1):
        if ws.cell(row=fila, column=col_F).value is None:
            continue

        for g_name, cols_sabor, cols_general in grupos:
            tiene_uno_en_sabores = any(
                _es_uno(ws.cell(row=fila, column=c).value)
                for c in cols_sabor
            )
            if tiene_uno_en_sabores:
                for gen_col in cols_general:
                    ws.cell(row=fila, column=gen_col).value = 1
                    total_actualizaciones += 1

    print(f"  ✓ Total de celdas de general marcadas: {total_actualizaciones}")
    return True
