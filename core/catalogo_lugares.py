"""
Actualizar catálogo de lugares (layout_places)
-----------------------------------------------
Compara el Maestro de lugares del cliente contra nuestro catálogo (layout_places)
y actualiza los campos que hayan cambiado:

  Cliente (Maestro)               →  Sistema (layout_places)
  ────────────────────────────────────────────────────────────
  BRANCHID                        →  Código Interno  (clave de búsqueda)
  NOMBRE DE LA TIENDA / Nombre Lugar →  Nombre Lugar
  ESTADO                          →  tags_ESTADO  +  tags_region_precios
  LATITUD                         →  Latitud
  LONGITUD                        →  Longitud
  STATUS OPERACIONES / Activo     →  Activo  (ACTIVO→1, INACTIVO→0, 1/0 pasan directo)

Cuando una fila es modificada se pone "UPDATE" en la columna Acción.
El resultado se guarda en carpeta_salida con el nombre del layout_places original.
"""

import os
import shutil
import openpyxl


# ── Columnas del archivo del cliente ─────────────────────────────────
# Las columnas marcadas como opcionales no son obligatorias; se usan si están presentes.
_COLS_CLIENTE = [
    "BRANCHID",
    "NOMBRE DE LA TIENDA",
    "Nombre Lugar",        # opcional: alternativa a NOMBRE DE LA TIENDA
    "ESTADO",
    "LATITUD",
    "LONGITUD",
    "STATUS OPERACIONES",  # opcional: fuente para columna Activo
    "Activo",              # opcional: alternativa a STATUS OPERACIONES
]

# Columnas que DEBEN existir (al menos una de cada grupo alternativo)
_COLS_CLIENTE_REQUERIDAS = ["BRANCHID", "ESTADO", "LATITUD", "LONGITUD"]
_COLS_NOMBRE_TIENDA = ["NOMBRE DE LA TIENDA", "Nombre Lugar"]  # al menos una
_COLS_ACTIVO = ["STATUS OPERACIONES", "Activo"]                # ambas opcionales

# ── Columnas del sistema (layout_places) ─────────────────────────────
_COLS_SISTEMA = [
    "Código Interno",
    "Nombre Lugar",
    "Cadena",
    "tags_ESTADO",
    "tags_region_precios",
    "Latitud",
    "Longitud",
    "Activo",
    "Acción",
]

# ── Mapeo: columna cliente → columna(s) sistema ───────────────────────
# El valor puede ser un str o una lista de str cuando se actualiza en varios lugares.
# Nota: "NOMBRE DE LA TIENDA" → "Nombre Lugar" se maneja por separado porque
# el nombre canónico se construye como: Código Interno + Cadena (catálogo) + Nombre Tienda.
_MAPEO = {
    "ESTADO":  ["tags_ESTADO", "tags_region_precios"],
    "LATITUD": "Latitud",
    "LONGITUD": "Longitud",
}

# ── Nombres de hoja esperados ────────────────────────────────────────
_HOJA_CLIENTE = "BD"
_HOJA_SISTEMA = "Lugares"

# ── Límite de registros por archivo de salida ────────────────────────
_LIMITE_REGISTROS = 25_000


# ══════════════════════════════════════════════════════════════════════
# Utilidades internas
# ══════════════════════════════════════════════════════════════════════

def _encontrar_fila_encabezados(ws, columnas_clave: list[str], max_filas: int = 40) -> int:
    """
    Devuelve el número de fila (1-based) donde están los encabezados.
    Busca la fila con el mayor número de coincidencias con columnas_clave.
    Esto maneja archivos del cliente que tienen filas extra arriba del encabezado.
    """
    clave_lower = {c.strip().lower() for c in columnas_clave}
    mejor_fila = 1
    mejor_hits = 0

    for fila_idx, fila in enumerate(
            ws.iter_rows(min_row=1, max_row=max_filas, values_only=True), start=1):
        vals = {str(v).strip().lower() for v in fila if v is not None}
        hits = len(clave_lower & vals)
        if hits > mejor_hits:
            mejor_hits = hits
            mejor_fila = fila_idx

    return mejor_fila


def _mapear_indices(ws, fila_enc: int, nombres: set) -> dict[str, int]:
    """
    Dado el número de fila de encabezados, devuelve {nombre_columna: col_idx_1based}
    para los nombres que se encuentren en esa fila.
    """
    fila = list(ws.iter_rows(
        min_row=fila_enc, max_row=fila_enc, values_only=True))[0]
    mapa = {}
    for col_idx, v in enumerate(fila, start=1):
        if v is None:
            continue
        nombre = str(v).strip()
        if nombre in nombres:
            mapa[nombre] = col_idx
    return mapa


def _normalizar_palabra(palabra: str) -> str:
    """
    Normaliza una palabra para comparación de duplicados (maneja singular/plural).
    Quita la 'S' final si la palabra tiene más de 3 letras.
    """
    palabra = palabra.upper().strip()
    if palabra.endswith('S') and len(palabra) > 3:
        return palabra[:-1]
    return palabra


def _palabras_similares(p1: str, p2: str) -> bool:
    """Verdadero si dos palabras son iguales o una es singular/plural de la otra."""
    return _normalizar_palabra(p1) == _normalizar_palabra(p2)


def _eliminar_duplicados_consecutivos(texto: str) -> str:
    """
    Elimina secuencias de palabras duplicadas consecutivas.
    Prueba longitudes desde la más larga posible hasta 1 para cubrir
    cadenas de cualquier número de palabras.
    Ejemplos:
      'SANTA CRUZ SANTA CRUZ CALVARIO'                         →  'SANTA CRUZ CALVARIO'
      'WELTON WELTON PROGRESO'                                 →  'WELTON PROGRESO'
      'FARMACIAS LA MAS BARATA FARMACIAS LA MAS BARATA MACRO'  →  'FARMACIAS LA MAS BARATA MACRO'
    """
    palabras = texto.split()
    n = len(palabras)
    if n < 2:
        return texto

    resultado = []
    i = 0
    while i < n:
        encontrado = False
        max_longitud = (n - i) // 2  # mayor secuencia posible desde la posición actual
        for longitud in range(max_longitud, 0, -1):
            if i + longitud * 2 <= n:
                seq1 = palabras[i:i + longitud]
                seq2 = palabras[i + longitud:i + longitud * 2]
                if all(_palabras_similares(seq1[j], seq2[j]) for j in range(longitud)):
                    resultado.extend(seq1)
                    i += longitud * 2
                    encontrado = True
                    break
        if not encontrado:
            resultado.append(palabras[i])
            i += 1

    return " ".join(resultado)


# Palabras que se ignoran al calcular siglas de una cadena
_STOP_SIGLAS = {
    'DE', 'DEL', 'LA', 'LAS', 'EL', 'LOS', 'Y', 'E', 'A',
    'EN', 'CON', 'SAN', 'SANTA', 'SANTO',
}


def _abreviatura_cadena(cadena: str) -> str:
    """
    Genera las siglas de una cadena tomando la primera letra de cada palabra
    significativa (omite artículos, preposiciones y conjunciones comunes).
    Retorna '' si el resultado tiene menos de 2 caracteres (muy genérico).
    Ejemplos:
      'FARMACIAS DEL AHORRO'  →  'FA'
      'FARMACIA GUADALAJARA'  →  'FG'
      'OXXO'                  →  ''   (una sola inicial, no se usa)
    """
    palabras = cadena.upper().split()
    siglas = "".join(p[0] for p in palabras if p and p not in _STOP_SIGLAS)
    return siglas if len(siglas) >= 2 else ""


def _quitar_siglas_cadena(nombre_tienda: str, cadena: str) -> str:
    """
    Si el nombre de tienda empieza con el nombre completo de la cadena o con
    sus siglas (como palabra(s) completa(s) al inicio), los elimina.
    Siempre compara en mayúsculas para no depender del formato del maestro.
    Ejemplos:
      ('FA JOSEFA',               'FARMACIAS DEL AHORRO')  →  'JOSEFA'
      ('FG LOMAS',                'FARMACIA GUADALAJARA')   →  'LOMAS'
      ('FARMACIAS LA MAS BARATA MACRO', 'FARMACIAS LA MAS BARATA')  →  'MACRO'
      ('JOSEFA',                  'FARMACIAS DEL AHORRO')   →  'JOSEFA'  (sin cambio)
    """
    if not cadena or not nombre_tienda:
        return nombre_tienda
    nombre_upper = nombre_tienda.upper()
    cadena_upper = cadena.upper()

    # 1. Verificar si el nombre empieza con el nombre completo de la cadena
    if nombre_upper == cadena_upper:
        return ""
    if nombre_upper.startswith(cadena_upper + " "):
        return nombre_tienda[len(cadena):].strip()

    # 2. Verificar con siglas de la cadena
    siglas = _abreviatura_cadena(cadena)
    if not siglas:
        return nombre_tienda
    if nombre_upper == siglas:
        return ""
    if nombre_upper.startswith(siglas + " "):
        return nombre_tienda[len(siglas):].strip()

    return nombre_tienda


def _normalizar(valor) -> str:
    """
    Convierte un valor de celda a str limpio y canónico para comparación.
    - None              → ""
    - int               → "12345"
    - float entero      → "12345"   (12345.0 → "12345")
    - str "12345.0"     → "12345"   (texto que parece entero → forma canónica)
    - str "12345"       → "12345"   (stripeado de espacios e invisibles)
    Garantiza que el mismo código almacenado como número o como texto siempre
    produzca el mismo resultado.
    """
    if valor is None:
        return ""
    if isinstance(valor, int):
        return str(valor)
    if isinstance(valor, float):
        if valor == int(valor):
            return str(int(valor))
        return str(valor)
    # --- tipo str u otro ---
    s = str(valor)
    # Eliminar caracteres invisibles comunes (espacio no rompible, zero-width,
    # BOM, soft-hyphen) que podrían venir de copiar/pegar o de Excel.
    for ch in ('\xa0', '\u200b', '\u200c', '\u200d', '\ufeff', '\u00ad'):
        s = s.replace(ch, '')
    s = s.strip()
    # Si el texto parece un número entero ("12345" o "12345.0"), normalizarlo
    # a la forma canónica para que coincida con valores numéricos del otro archivo.
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except (ValueError, OverflowError):
        pass
    return s


# ══════════════════════════════════════════════════════════════════════
# Función principal
# ══════════════════════════════════════════════════════════════════════

def actualizar_catalogo_lugares(
        ruta_maestro: str,
        ruta_layout_places: str,
        carpeta_salida: str) -> bool:
    """
    Actualiza layout_places con los datos del Maestro de lugares.

    Args:
        ruta_maestro:       Ruta al archivo Maestro de lugares (cliente).
        ruta_layout_places: Ruta al archivo layout_places (nuestro catálogo).
        carpeta_salida:     Carpeta donde se guardará el resultado.

    Returns:
        True si el proceso terminó correctamente, False si hubo un error grave.
    """

    print("=" * 55)
    print("--- Actualizar Catálogo de Lugares ---")
    print("=" * 55)
    print(f"  Maestro : {os.path.basename(ruta_maestro)}")
    print(f"  Catálogo: {os.path.basename(ruta_layout_places)}")

    # ── 1. Cargar Maestro de lugares (cliente) ────────────────────────
    try:
        wb_cliente = openpyxl.load_workbook(ruta_maestro, read_only=True, data_only=True, keep_links=False)
    except PermissionError:
        raise RuntimeError("El Maestro de lugares está abierto en Excel. Ciérralo e intenta de nuevo.")
    except Exception as exc:
        raise RuntimeError(f"No se pudo abrir el Maestro de lugares: {exc}")

    if _HOJA_CLIENTE not in wb_cliente.sheetnames:
        available = ", ".join(wb_cliente.sheetnames)
        wb_cliente.close()
        raise RuntimeError(f"No se encontró la hoja '{_HOJA_CLIENTE}' en el Maestro de lugares.\n"
                           f"Hojas disponibles: {available}")
    ws_cliente = wb_cliente[_HOJA_CLIENTE]

    # Detectar fila de encabezados en el archivo del cliente
    fila_enc_cliente = _encontrar_fila_encabezados(ws_cliente, _COLS_CLIENTE)
    print(f"  Fila de encabezados (Maestro): {fila_enc_cliente}")

    mapa_cliente = _mapear_indices(ws_cliente, fila_enc_cliente, set(_COLS_CLIENTE))

    # Verificar columnas mínimas requeridas
    faltantes = [c for c in _COLS_CLIENTE_REQUERIDAS if c not in mapa_cliente]
    if faltantes:
        wb_cliente.close()
        raise RuntimeError(f"Columnas no encontradas en Maestro de lugares: {faltantes}\n"
                           f"Encabezados detectados en fila {fila_enc_cliente}.")
    if not any(c in mapa_cliente for c in _COLS_NOMBRE_TIENDA):
        wb_cliente.close()
        raise RuntimeError(f"No se encontró ninguna columna de nombre de tienda "
                           f"({' / '.join(_COLS_NOMBRE_TIENDA)}) en el Maestro.\n"
                           f"Encabezados detectados en fila {fila_enc_cliente}.")

    # Construir diccionario BRANCHID → {col_cliente: valor_str}
    datos_cliente: dict[str, dict[str, str]] = {}
    col_branch = mapa_cliente["BRANCHID"]

    for fila in ws_cliente.iter_rows(min_row=fila_enc_cliente + 1, values_only=True):
        branch_raw = fila[col_branch - 1]
        if branch_raw is None:
            continue
        branch_id = _normalizar(branch_raw)
        if not branch_id:
            continue
        datos_cliente[branch_id] = {
            col: _normalizar(fila[idx - 1])
            for col, idx in mapa_cliente.items()
        }

    wb_cliente.close()
    print(f"  Registros en Maestro: {len(datos_cliente)}")

    # ── 2. Copiar layout_places a la carpeta de salida ────────────────
    ext_salida = os.path.splitext(ruta_layout_places)[1]
    nombre_salida = f"layout_places_actualizado{ext_salida}"
    ruta_salida = os.path.join(carpeta_salida, nombre_salida)
    try:
        shutil.copy2(ruta_layout_places, ruta_salida)
    except Exception as exc:
        raise RuntimeError(f"No se pudo copiar layout_places: {exc}")

    # ── 3. Abrir la copia y actualizar ───────────────────────────────
    try:
        wb_sistema = openpyxl.load_workbook(ruta_salida, keep_links=False)
    except Exception as exc:
        raise RuntimeError(f"No se pudo abrir layout_places: {exc}")

    if _HOJA_SISTEMA not in wb_sistema.sheetnames:
        available = ", ".join(wb_sistema.sheetnames)
        wb_sistema.close()
        raise RuntimeError(f"No se encontró la hoja '{_HOJA_SISTEMA}' en layout_places.\n"
                           f"Hojas disponibles: {available}")
    ws_sistema = wb_sistema[_HOJA_SISTEMA]

    fila_enc_sistema = _encontrar_fila_encabezados(ws_sistema, _COLS_SISTEMA)
    print(f"  Fila de encabezados (layout_places): {fila_enc_sistema}")

    mapa_sistema = _mapear_indices(ws_sistema, fila_enc_sistema, set(_COLS_SISTEMA))

    # Verificar columnas críticas del sistema
    for col_req in ("Código Interno", "Acción"):
        if col_req not in mapa_sistema:
            wb_sistema.close()
            raise RuntimeError(f"No se encontró la columna '{col_req}' en layout_places.\n"
                               f"Encabezados detectados en fila {fila_enc_sistema}.")

    col_codigo_int = mapa_sistema["Código Interno"]
    col_accion     = mapa_sistema["Acción"]

    # ── 3b. Detectar códigos del Maestro que NO están en el catálogo ──
    codigos_catalogo: set[str] = set()
    for fila in ws_sistema.iter_rows(min_row=fila_enc_sistema + 1, values_only=True):
        val = fila[col_codigo_int - 1]
        if val is not None:
            codigos_catalogo.add(_normalizar(val))

    # Reabrir Maestro para extraer filas completas de los que faltan
    wb_cliente2 = openpyxl.load_workbook(ruta_maestro, read_only=True, data_only=True, keep_links=False)
    ws_cliente2 = wb_cliente2[_HOJA_CLIENTE]

    faltantes_maestro = []
    encabezados_maestro = None
    for idx_fila, fila in enumerate(
            ws_cliente2.iter_rows(min_row=fila_enc_cliente, values_only=True), start=fila_enc_cliente):
        if idx_fila == fila_enc_cliente:
            encabezados_maestro = fila
            continue
        branch_raw = fila[col_branch - 1]
        if branch_raw is None:
            continue
        branch_id = _normalizar(branch_raw)
        if branch_id and branch_id not in codigos_catalogo:
            faltantes_maestro.append(fila)

    wb_cliente2.close()

    if faltantes_maestro:
        wb_falt = openpyxl.Workbook()
        ws_falt = wb_falt.active
        ws_falt.title = "No encontrados"
        if encabezados_maestro:
            ws_falt.append(list(encabezados_maestro))
        for fila_data in faltantes_maestro:
            ws_falt.append(list(fila_data))
        ruta_faltantes = os.path.join(carpeta_salida, "codigos_no_encontrados_en_catalogo.xlsx")
        wb_falt.save(ruta_faltantes)
        wb_falt.close()
        print(f"  Códigos del Maestro no encontrados en catálogo: {len(faltantes_maestro)}")
        print(f"  Archivo generado: {os.path.basename(ruta_faltantes)}")

    actualizados   = 0
    sin_cambios    = 0
    no_encontrados = 0

    # Pre-cargar todas las filas de datos como objetos de celda en memoria
    # Esto evita llamadas repetidas a ws.cell(row, col) que son lentas
    todas_las_filas = list(ws_sistema.iter_rows(min_row=fila_enc_sistema + 1))

    for fila_cells in todas_las_filas:

        codigo_raw = fila_cells[col_codigo_int - 1].value
        if codigo_raw is None:
            continue
        codigo_str = _normalizar(codigo_raw)
        if not codigo_str:
            continue

        if codigo_str not in datos_cliente:
            no_encontrados += 1
            continue

        datos = datos_cliente[codigo_str]
        fila_modificada = False

        # Recorrer el mapeo cliente → sistema
        for col_cliente, destino in _MAPEO.items():
            nuevo_val_str = datos.get(col_cliente, "")
            if not nuevo_val_str:
                # No sobreescribir con vacío
                continue

            # destino puede ser str o lista de str
            destinos = [destino] if isinstance(destino, str) else destino

            for col_sis in destinos:
                if col_sis not in mapa_sistema:
                    continue

                cell = fila_cells[mapa_sistema[col_sis] - 1]
                actual_str = _normalizar(cell.value)

                if nuevo_val_str == actual_str:
                    continue  # Sin cambio

                # Intentar preservar el tipo numérico para Latitud/Longitud
                if col_sis in ("Latitud", "Longitud"):
                    try:
                        cell.value = float(nuevo_val_str)
                    except ValueError:
                        cell.value = nuevo_val_str
                else:
                    cell.value = nuevo_val_str

                fila_modificada = True

        # ── Actualizar Nombre Lugar (lógica especial) ────────────────────
        # El nombre canónico es: Código Interno + Cadena (del catálogo) + Nombre Tienda (del Maestro)
        # La Cadena se toma siempre del catálogo para no usar el valor incorrecto del Maestro.
        # El maestro puede traer el nombre en "NOMBRE DE LA TIENDA" o en "Nombre Lugar".
        nombre_tienda_nuevo = datos.get("NOMBRE DE LA TIENDA") or datos.get("Nombre Lugar", "")
        if nombre_tienda_nuevo and "Nombre Lugar" in mapa_sistema:
            cadena = ""
            if "Cadena" in mapa_sistema:
                cadena = _normalizar(fila_cells[mapa_sistema["Cadena"] - 1].value)
            # Quitar siglas de la cadena si el maestro las incrustó en el nombre
            # Ej: 'FA JOSEFA' con cadena 'FARMACIAS DEL AHORRO' → 'JOSEFA'
            nombre_tienda_limpio = _quitar_siglas_cadena(nombre_tienda_nuevo, cadena)
            partes = [p for p in [codigo_str, cadena, nombre_tienda_limpio] if p]
            nombre_nuevo = _eliminar_duplicados_consecutivos(" ".join(partes))
            nombre_nuevo = " ".join(nombre_nuevo.split())  # limpiar espacios dobles
            if len(nombre_nuevo) > 60:
                nombre_nuevo = nombre_nuevo[:60]

            cell_nombre = fila_cells[mapa_sistema["Nombre Lugar"] - 1]
            actual_nombre = (cell_nombre.value or "").strip()

            if nombre_nuevo != actual_nombre:
                cell_nombre.value = nombre_nuevo
                fila_modificada = True

        # ── Actualizar Activo ────────────────────────────────────────────
        # El maestro puede tener la columna "STATUS OPERACIONES" o "Activo".
        # Valores aceptados: ACTIVO → 1, INACTIVO → 0, 1 → 1, 0 → 0.
        if "Activo" in mapa_sistema:
            estado_raw = datos.get("STATUS OPERACIONES") or datos.get("Activo", "")
            if estado_raw:
                estado_upper = estado_raw.upper()
                if estado_upper == "ACTIVO":
                    nuevo_activo = 1
                elif estado_upper == "INACTIVO":
                    nuevo_activo = 0
                elif estado_upper in ("1", "0"):
                    nuevo_activo = int(estado_upper)
                else:
                    nuevo_activo = None

                if nuevo_activo is not None:
                    cell_activo = fila_cells[mapa_sistema["Activo"] - 1]
                    if _normalizar(nuevo_activo) != _normalizar(cell_activo.value):
                        cell_activo.value = nuevo_activo
                        fila_modificada = True

        if fila_modificada:
            fila_cells[col_accion - 1].value = "UPDATE"
            actualizados += 1
        else:
            sin_cambios += 1

    # ── 4. Guardar ────────────────────────────────────────────────────
    try:
        wb_sistema.save(ruta_salida)
    except PermissionError:
        wb_sistema.close()
        raise RuntimeError("El archivo de salida está abierto en Excel. Ciérralo e intenta de nuevo.")
    except Exception as exc:
        wb_sistema.close()
        raise RuntimeError(f"No se pudo guardar el archivo: {exc}")

    wb_sistema.close()

    # ── 5. Dividir en partes si supera el límite ──────────────────────
    total_filas_datos = len(todas_las_filas)
    if total_filas_datos > _LIMITE_REGISTROS:
        print(f"  El catálogo tiene {total_filas_datos:,} registros → "
              f"dividiendo en partes de {_LIMITE_REGISTROS:,}…")

        # Extraer valores de las filas de datos (ya cargadas en memoria como Cell objects)
        filas_datos = [[cell.value for cell in row] for row in todas_las_filas]

        ext = os.path.splitext(ruta_salida)[1]
        nombre_base = "layout_places_actualizado"
        num_partes = (total_filas_datos + _LIMITE_REGISTROS - 1) // _LIMITE_REGISTROS

        for i in range(num_partes):
            chunk = filas_datos[i * _LIMITE_REGISTROS:(i + 1) * _LIMITE_REGISTROS]
            ruta_p = os.path.join(carpeta_salida, f"{nombre_base}_parte_{i + 1}{ext}")
            # Copiar el archivo actualizado para preservar todo el formato y propiedades
            shutil.copy2(ruta_salida, ruta_p)
            wb_p = openpyxl.load_workbook(ruta_p, keep_links=False)
            ws_p = wb_p[_HOJA_SISTEMA]
            # Eliminar filas de datos anteriores; conservar encabezados
            filas_en_copia = ws_p.max_row
            if filas_en_copia > fila_enc_sistema:
                ws_p.delete_rows(fila_enc_sistema + 1, filas_en_copia - fila_enc_sistema)
            for fila in chunk:
                ws_p.append(fila)
            wb_p.save(ruta_p)
            wb_p.close()
            print(f"    Parte {i + 1}/{num_partes}: {len(chunk):,} registros "
                  f"→ {os.path.basename(ruta_p)}")

        # Eliminar el archivo unificado; queda reemplazado por las partes
        os.remove(ruta_salida)
        print(f"  División completada: {num_partes} archivos generados.")

    return True

