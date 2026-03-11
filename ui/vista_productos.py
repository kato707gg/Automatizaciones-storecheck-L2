"""
Vista – Editor de Productos de Sams
Carga el archivo 'Productos de Sams.xlsx' y permite editar las 3 listas
(Generales, SAMS, Competencia) sin necesidad de abrir Excel.
"""

import os
import openpyxl
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QMessageBox,
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont

# Palabras clave para detectar cada grupo por su encabezado (fila 1 del Excel)
KWORD_NORMAL = "captura normal"
KWORD_COMP   = "competencia"
KWORD_SAMS   = "sams"


# ── Tabla editable ───────────────────────────────────────────────────
class TablaProductos(QWidget):
    """Tabla editable de 2 columnas (id sto | nombre) con botones + / −."""

    HEADER_COLOR = {
        KWORD_NORMAL: "#1570C0",   # azul oscuro
        KWORD_COMP:   "#C07C15",   # azul oscuro
        KWORD_SAMS:   "#2E2F7D",   # verde oscuro
    }

    def __init__(self, titulo: str, col_nombre: str, grupo: str,
                 filas: list, parent=None):
        super().__init__(parent)
        col_bg = self.HEADER_COLOR.get(grupo, "#1565C0")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        # Título de sección
        lbl = QLabel(titulo)
        lbl.setFont(QFont("Segoe UI", 14, QFont.Bold))
        lbl.setStyleSheet("color: #0098C4;")
        layout.addWidget(lbl)

        # Tabla
        self.tabla = QTableWidget()
        self.tabla.setColumnCount(2)
        self.tabla.setHorizontalHeaderLabels(["id sto", col_nombre])
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.tabla.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabla.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tabla.setAlternatingRowColors(False)
        self.tabla.verticalHeader().setVisible(False)
        self.tabla.setEditTriggers(
            QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed
        )
        self.tabla.setStyleSheet(f"""
            QHeaderView::section {{
                background-color: {col_bg};
                color: #FFFFFF;
                font-weight: bold;
                font-size: 11px;
                padding: 5px 4px;
                border: none;
            }}
            QTableWidget {{
                border: 1px solid #DADADA;
                border-radius: 4px;
                gridline-color: #EEEEEE;
                font-size: 11px;
            }}
            QTableWidget::item:selected {{
                background-color: #BBDEFB;
                color: #000000;
            }}

        """)
        self._cargar_filas(filas)
        layout.addWidget(self.tabla, 1)

        # Botones + / −
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        btn_add = QPushButton("+ Agregar")
        btn_del = QPushButton("− Eliminar")
        for btn in (btn_add, btn_del):
            btn.setFixedHeight(28)
            btn.setFont(QFont("Segoe UI", 9))
            btn.setCursor(Qt.PointingHandCursor)
        btn_add.setStyleSheet("""
            QPushButton { background:#E3F2FD; color:#1565C0;
                border:1px solid #1565C0; border-radius:5px; padding:0 10px; }
            QPushButton:hover { background:#BBDEFB; }
        """)
        btn_del.setStyleSheet("""
            QPushButton { background:#FFEBEE; color:#B71C1C;
                border:1px solid #B71C1C; border-radius:5px; padding:0 10px; }
            QPushButton:hover { background:#FFCDD2; }
        """)
        btn_add.clicked.connect(self._agregar_fila)
        btn_del.clicked.connect(self._eliminar_fila)
        btn_row.addWidget(btn_add)
        btn_row.addWidget(btn_del)
        btn_row.addStretch()
        layout.addLayout(btn_row)

    # ── Datos ────────────────────────────────────────────────────────
    def _cargar_filas(self, filas: list):
        self.tabla.setRowCount(0)
        for id_val, nom_val in filas:
            r = self.tabla.rowCount()
            self.tabla.insertRow(r)
            self.tabla.setItem(r, 0, QTableWidgetItem(
                "" if id_val is None else str(id_val)))
            self.tabla.setItem(r, 1, QTableWidgetItem(
                "" if nom_val is None else str(nom_val)))

    def _agregar_fila(self):
        r = self.tabla.rowCount()
        self.tabla.insertRow(r)
        self.tabla.setItem(r, 0, QTableWidgetItem(""))
        self.tabla.setItem(r, 1, QTableWidgetItem(""))
        self.tabla.scrollToBottom()
        self.tabla.setCurrentCell(r, 0)
        self.tabla.editItem(self.tabla.item(r, 0))

    def _eliminar_fila(self):
        rows = sorted(
            {i.row() for i in self.tabla.selectedIndexes()}, reverse=True
        )
        for r in rows:
            self.tabla.removeRow(r)

    def obtener_datos(self) -> list:
        """Devuelve [(id, nombre), ...] con los valores actuales."""
        datos = []
        for r in range(self.tabla.rowCount()):
            id_item  = self.tabla.item(r, 0)
            nom_item = self.tabla.item(r, 1)
            id_txt   = id_item.text().strip()  if id_item  else ""
            nom_txt  = nom_item.text().strip() if nom_item else ""
            if not id_txt and not nom_txt:
                continue
            try:
                id_val = int(id_txt)
            except ValueError:
                id_val = id_txt or None
            datos.append((id_val, nom_txt or None))
        return datos


# ── Vista principal ──────────────────────────────────────────────────
class VistaProductos(QWidget):
    def __init__(self, back_cb=None, parent=None):
        super().__init__(parent)
        self._archivo  = None
        self._d_normal = []   # datos sin modificar
        self._t_sams   = None
        self._t_comp   = None
        self.setStyleSheet("VistaProductos { background-color: #FFFFFF; }")

        outer = QVBoxLayout(self)
        outer.setContentsMargins(40, 36, 40, 36)
        outer.setSpacing(0)

        # ── Botón volver ─────────────────────────────────────────────
        btn_back = QPushButton("← Volver")
        btn_back.setFixedSize(120, 36)
        btn_back.setFont(QFont("Segoe UI", 10))
        btn_back.setCursor(Qt.PointingHandCursor)
        btn_back.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #0098C4;
                border: 1.5px solid #0098C4;
                border-radius: 18px;
            }
            QPushButton:hover { background-color: #E8F7FC; }
        """)
        if back_cb:
            btn_back.clicked.connect(back_cb)
        outer.addWidget(btn_back, 0, Qt.AlignLeft)
        outer.addSpacing(24)

        # ── Contenedor de las 3 tablas ───────────────────────────────
        self._tablas_layout = QHBoxLayout()
        self._tablas_layout.setSpacing(30)
        outer.addLayout(self._tablas_layout, 1)
        outer.addSpacing(24)

        # ── Botón guardar ────────────────────────────────────────────
        self._btn_guardar = QPushButton("Guardar cambios")
        self._btn_guardar.setFixedSize(240, 50)
        self._btn_guardar.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self._btn_guardar.setCursor(Qt.PointingHandCursor)
        self._btn_guardar.setStyleSheet("""
            QPushButton {
                background-color: #0098C4;
                color: #FFFFFF;
                border: none;
                border-radius: 25px;
            }
            QPushButton:hover   { background-color: #007BA3; }
            QPushButton:pressed { background-color: #006080; }
            QPushButton:disabled { background-color: #CCCCCC; }
        """)
        self._btn_guardar.clicked.connect(self._guardar)
        outer.addWidget(self._btn_guardar, 0, Qt.AlignHCenter)

        self._cargar_datos()

    # ── Carga de datos ───────────────────────────────────────────────
    def _buscar_archivo(self):
        # Los xlsx de catalogación viven en core/catalogacion/data/
        ruta_base = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "core", "catalogacion", "data")
        for nombre in os.listdir(ruta_base):
            if "productos de sams" in nombre.lower() and \
               nombre.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                return os.path.join(ruta_base, nombre)
        return None

    def _cargar_datos(self):
        self._archivo = self._buscar_archivo()
        if not self._archivo:
            lbl = QLabel(
                "No se encontró el archivo 'Productos de Sams'.\n"
                "Verifica que esté en la misma carpeta que la aplicación."
            )
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet("color: #888; font-size: 13px;")
            self._tablas_layout.addWidget(lbl)
            self._btn_guardar.setEnabled(False)
            return

        wb = openpyxl.load_workbook(self._archivo, data_only=True, read_only=True)
        rows_all = [list(f) for f in wb.active.iter_rows(values_only=True)]
        wb.close()

        if not rows_all:
            return

        header_row = rows_all[0]

        # Detectar columnas por texto del encabezado
        col_id_n = col_nom_n = nombre_n = None
        col_id_c = col_nom_c = nombre_c = None
        col_id_s = col_nom_s = nombre_s = None

        for idx, h in enumerate(header_row):
            if h is None:
                continue
            hl = str(h).lower()
            if KWORD_NORMAL in hl:
                col_nom_n, col_id_n, nombre_n = idx, idx - 1, str(h)
            elif KWORD_COMP in hl:
                col_nom_c, col_id_c, nombre_c = idx, idx - 1, str(h)
            elif KWORD_SAMS in hl:
                col_nom_s, col_id_s, nombre_s = idx, idx - 1, str(h)

        def extraer(col_id, col_nom):
            result = []
            for fila in rows_all[1:]:
                iv = fila[col_id]  if col_id  is not None and col_id  < len(fila) else None
                nv = fila[col_nom] if col_nom is not None and col_nom < len(fila) else None
                if iv is not None or nv is not None:
                    result.append((iv, nv))
            return result

        self._d_normal = extraer(col_id_n, col_nom_n) if col_id_n is not None else []

        self._t_sams = TablaProductos(
            "Productos SAMS",
            nombre_s or "Productos en SAMS",
            KWORD_SAMS,
            extraer(col_id_s, col_nom_s) if col_id_s is not None else [],
        )
        self._t_comp = TablaProductos(
            "Productos Competencia",
            nombre_c or "Productos Competencia",
            KWORD_COMP,
            extraer(col_id_c, col_nom_c) if col_id_c is not None else [],
        )

        self._tablas_layout.addWidget(self._t_sams)
        self._tablas_layout.addWidget(self._t_comp)

    # ── Guardar ──────────────────────────────────────────────────────
    def _guardar(self):
        if not self._archivo:
            return

        d_n = self._d_normal  # sin modificar
        d_s = self._t_sams.obtener_datos()   if self._t_sams   else []
        d_c = self._t_comp.obtener_datos()   if self._t_comp   else []

        try:
            wb = openpyxl.load_workbook(self._archivo)
            ws = wb.active

            # Limpiar desde fila 2 (preserva encabezados en fila 1)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None

            # Escribir grupos en sus columnas originales
            max_filas = max(len(d_n), len(d_c), len(d_s), 1)
            for i in range(max_filas):
                fila = i + 2
                if i < len(d_n):
                    ws.cell(row=fila, column=1).value = d_n[i][0]  # A
                    ws.cell(row=fila, column=2).value = d_n[i][1]  # B
                if i < len(d_c):
                    ws.cell(row=fila, column=3).value = d_c[i][0]  # C
                    ws.cell(row=fila, column=4).value = d_c[i][1]  # D
                if i < len(d_s):
                    ws.cell(row=fila, column=5).value = d_s[i][0]  # E
                    ws.cell(row=fila, column=6).value = d_s[i][1]  # F

            wb.save(self._archivo)
            wb.close()
            QMessageBox.information(self, "Guardado", "✓ Cambios guardados correctamente.")

        except PermissionError:
            QMessageBox.critical(
                self, "Error",
                "No se pudo guardar porque el archivo está abierto en Excel.\n"
                "Ciérralo e intenta de nuevo."
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al guardar:\n{str(e)}")
